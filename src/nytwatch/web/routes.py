from __future__ import annotations

import json
import logging
import re
import threading
from io import BytesIO
from datetime import datetime
from pathlib import Path
from typing import Optional

from fastapi import APIRouter, Request, WebSocket, WebSocketDisconnect
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse, StreamingResponse
from fastapi.templating import Jinja2Templates

from nytwatch.database import Database
from nytwatch.models import BatchStatus, FindingStatus, now_iso

logger = logging.getLogger(__name__)

router = APIRouter()

templates = Jinja2Templates(
    directory=str(Path(__file__).parent / "templates")
)

# ── Template globals ─────────────────────────────────────────────────────────
# Callables receive `request` so they always reflect the current active project
# without requiring every route to explicitly pass these values.

def _active_project_name(request: "Request") -> str:
    """Return the short project name (repo folder) for the active project."""
    config = getattr(request.app.state, "config", None)
    if config and getattr(config, "repo_path", ""):
        return Path(config.repo_path).name
    return ""

def _active_config_path(request: "Request") -> str:
    return getattr(request.app.state, "config_path", "") or ""

def _active_repo_path(request: "Request") -> str:
    config = getattr(request.app.state, "config", None)
    return getattr(config, "repo_path", "") if config else ""

templates.env.globals["active_project_name"] = _active_project_name
templates.env.globals["active_config_path"] = _active_config_path
templates.env.globals["active_repo_path"] = _active_repo_path


def get_db(request: Request) -> Optional[Database]:
    return request.app.state.db


def get_config(request: Request):
    return request.app.state.config


# --- Dashboard ---

def _infer_source_dirs(systems: list[dict], source_dir_paths: set[str]) -> list[dict]:
    """For systems with an empty source_dir, infer it by longest-prefix match against
    the known source directories.  Returns the full list (modified in place for matches)."""
    # Normalise every source dir to "path/with/trailing/slash"
    sd_normed = {sd: sd.replace("\\", "/").rstrip("/") + "/" for sd in source_dir_paths}

    result = []
    for s in systems:
        if s.get("source_dir"):
            result.append(s)
            continue
        best_sd, best_len = "", 0
        for path in s.get("paths", []):
            p = path.replace("\\", "/")
            if not p.endswith("/"):
                p += "/"
            for sd, sd_norm in sd_normed.items():
                if p.startswith(sd_norm) and len(sd_norm) > best_len:
                    best_sd, best_len = sd, len(sd_norm)
        result.append(dict(s, source_dir=best_sd) if best_sd else s)
    return result


@router.get("/", response_class=HTMLResponse)
async def dashboard(request: Request):
    db = get_db(request)
    config = get_config(request)

    # First-run: no project configured yet → redirect to setup wizard.
    # Only redirect when there is genuinely no active project (no repo_path).
    # A configured project with no systems yet is valid and should show the dashboard.
    if not config.repo_path:
        return RedirectResponse(url="/settings?setup=1")

    db_systems = db.list_systems()
    all_source_dirs = db.list_source_dirs()
    source_dir_map = {d["path"]: d["source_type"] for d in all_source_dirs}

    # ── One-time repair: assign source_dir to systems that are missing it ──────
    needs_repair = any(not s.get("source_dir") for s in db_systems)
    if needs_repair and source_dir_map:
        repaired = _infer_source_dirs(db_systems, set(source_dir_map.keys()))
        db.replace_systems([
            {
                "name": s["name"],
                "source_dir": s.get("source_dir") or "",
                "paths": s.get("paths", []),
                "min_confidence": s.get("min_confidence"),
                "file_extensions": s.get("file_extensions"),
                "claude_fast_mode": s.get("claude_fast_mode"),
            }
            for s in repaired
        ])
        db_systems = db.list_systems()
        logger.info("Repaired source_dir for %d system(s)",
                    sum(1 for s in repaired if s.get("source_dir")))

    # ── Build grouped structure ───────────────────────────────────────────────
    sys_by_dir: dict[str, list] = {}
    for s in db_systems:
        sd = s.get("source_dir") or ""
        sys_by_dir.setdefault(sd, []).append({
            "id": s["id"],
            "name": s["name"],
            "source_dir": sd,
            "count": db.count_findings_for_path_prefixes(s["paths"]),
            "tracking_enabled": s.get("tracking_enabled", False),
        })

    # Sort groups: active before ignored, then alphabetically by source_dir name
    all_source_dirs = list(sys_by_dir.keys())
    all_source_dirs.sort(key=lambda sd: (
        1 if source_dir_map.get(sd) == "ignored" else 0,  # active first
        sd.lower(),
    ))

    grouped_systems = []
    for sd in all_source_dirs:
        systems_in_group = sorted(sys_by_dir[sd], key=lambda s: s["name"].lower())
        grouped_systems.append({
            "source_dir": sd,
            "ignored": source_dir_map.get(sd) == "ignored",
            "systems": systems_in_group,
        })

    has_grouping = any(g["source_dir"] for g in grouped_systems)

    # Flat list in the same order — used as JS index baseline
    systems = [s for g in grouped_systems for s in g["systems"]]

    stats = db.get_stats()
    batches = db.list_batches(limit=5)
    config_obj = get_config(request)
    pie_state = False
    watcher = getattr(request.app.state, "watcher", None)
    if watcher is not None and config_obj.repo_path:
        pie_state = watcher.get_pie_state(config_obj.repo_path)
    recent_sessions = db.list_sessions(project_dir=config_obj.repo_path or None, limit=3)
    armed_count = sum(1 for s in db_systems if s.get("tracking_enabled"))
    return templates.TemplateResponse(request, "dashboard.html", {
        "stats": stats,
        "batches": batches,
        "systems": systems,
        "grouped_systems": grouped_systems,
        "has_grouping": has_grouping,
        "pie_state": pie_state,
        "recent_sessions": recent_sessions,
        "armed_count": armed_count,
    })


# --- Auditor ---

@router.get("/auditor", response_class=HTMLResponse)
async def auditor_root(request: Request):
    return RedirectResponse(url="/auditor/findings")


@router.get("/auditor/findings", response_class=HTMLResponse)
async def auditor_findings(  # noqa: C901
    request: Request,
    status: Optional[str] = None,
    severity: Optional[str] = None,
    category: Optional[str] = None,
    confidence: Optional[str] = None,
    file_path: Optional[str] = None,
    source: Optional[str] = None,
    system: Optional[str] = None,
):
    db = get_db(request)
    if db is None:
        return RedirectResponse(url="/settings?setup=1")

    path_prefixes = None
    if system:
        sys_def = next((s for s in db.list_systems() if s["name"] == system), None)
        if sys_def:
            path_prefixes = sys_def["paths"]

    findings = db.list_findings(
        status=status,
        severity=severity,
        category=category,
        confidence=confidence,
        file_path=file_path,
        source=source,
        path_prefixes=path_prefixes,
    )
    approved_count = len(db.get_approved_findings())
    filters = {
        "status": status,
        "severity": severity,
        "category": category,
        "confidence": confidence,
        "file_path": file_path,
        "source": source,
        "system": system,
    }
    ignored_dirs = {
        row["path"].replace("\\", "/").rstrip("/")
        for row in db.list_source_dirs()
        if row["source_type"] == "ignored"
    }
    active_systems = [
        s for s in db.list_systems()
        if s.get("source_dir", "").replace("\\", "/").rstrip("/") not in ignored_dirs
    ]
    return templates.TemplateResponse(request, "auditor.html", {
        "active_tab": "findings",
        "findings": findings,
        "filters": filters,
        "approved_count": approved_count,
        "systems": active_systems,
    })


@router.get("/auditor/findings/export")
async def auditor_findings_export(request: Request):
    qs = str(request.url.query)
    url = f"/findings/export?{qs}" if qs else "/findings/export"
    return RedirectResponse(url=url)


@router.get("/auditor/findings/{finding_id}", response_class=HTMLResponse)
async def auditor_finding_detail(request: Request, finding_id: str):
    db = get_db(request)
    if db is None:
        return HTMLResponse("<h1>Finding not found</h1>", status_code=404)
    finding = db.get_finding(finding_id)
    if not finding:
        return HTMLResponse("<h1>Finding not found</h1>", status_code=404)
    return templates.TemplateResponse(request, "finding_detail.html", {
        "finding": finding,
    })


@router.get("/auditor/batches", response_class=HTMLResponse)
async def auditor_batches(request: Request):
    db = get_db(request)
    if db is None:
        return RedirectResponse(url="/settings?setup=1")
    batches = db.list_batches()
    return templates.TemplateResponse(request, "auditor.html", {
        "active_tab": "batches",
        "batches": batches,
    })


@router.get("/auditor/batches/{batch_id}", response_class=HTMLResponse)
async def auditor_batch_detail(request: Request, batch_id: str):
    db = get_db(request)
    if db is None:
        return HTMLResponse("<h1>Batch not found</h1>", status_code=404)
    batch = db.get_batch(batch_id)
    if not batch:
        return HTMLResponse("<h1>Batch not found</h1>", status_code=404)
    findings = [db.get_finding(fid) for fid in batch["finding_ids"]]
    findings = [f for f in findings if f]
    return templates.TemplateResponse(request, "batch_detail.html", {
        "batch": batch,
        "findings": findings,
    })


@router.get("/auditor/scans", response_class=HTMLResponse)
async def auditor_scans(request: Request):
    db = get_db(request)
    if db is None:
        return RedirectResponse(url="/settings?setup=1")
    scans = db.list_scans()
    log_counts = db.get_scan_log_counts()
    return templates.TemplateResponse(request, "auditor.html", {
        "active_tab": "scans",
        "scans": scans,
        "log_counts": log_counts,
    })


@router.get("/auditor/systems", response_class=HTMLResponse)
async def auditor_systems(request: Request):
    db = get_db(request)
    if db is None:
        return RedirectResponse(url="/settings?setup=1")

    db_systems = db.list_systems()
    all_source_dirs = db.list_source_dirs()
    source_dir_map = {d["path"]: d["source_type"] for d in all_source_dirs}

    needs_repair = any(not s.get("source_dir") for s in db_systems)
    if needs_repair and source_dir_map:
        repaired = _infer_source_dirs(db_systems, set(source_dir_map.keys()))
        db.replace_systems([
            {
                "name": s["name"],
                "source_dir": s.get("source_dir") or "",
                "paths": s.get("paths", []),
                "min_confidence": s.get("min_confidence"),
                "file_extensions": s.get("file_extensions"),
                "claude_fast_mode": s.get("claude_fast_mode"),
            }
            for s in repaired
        ])
        db_systems = db.list_systems()

    sys_by_dir: dict[str, list] = {}
    for s in db_systems:
        sd = s.get("source_dir") or ""
        sys_by_dir.setdefault(sd, []).append({
            "id": s["id"],
            "name": s["name"],
            "source_dir": sd,
            "count": db.count_findings_for_path_prefixes(s["paths"]),
            "tracking_enabled": s.get("tracking_enabled", False),
        })

    sorted_dirs = sorted(
        sys_by_dir.keys(),
        key=lambda sd: (1 if source_dir_map.get(sd) == "ignored" else 0, sd.lower()),
    )
    grouped_systems = []
    for sd in sorted_dirs:
        systems_in_group = sorted(sys_by_dir[sd], key=lambda s: s["name"].lower())
        grouped_systems.append({
            "source_dir": sd,
            "ignored": source_dir_map.get(sd) == "ignored",
            "systems": systems_in_group,
        })

    has_grouping = any(g["source_dir"] for g in grouped_systems)
    systems = [s for g in grouped_systems for s in g["systems"]]

    return templates.TemplateResponse(request, "auditor.html", {
        "active_tab": "systems",
        "systems": systems,
        "grouped_systems": grouped_systems,
        "has_grouping": has_grouping,
    })


# --- Findings ---

@router.get("/findings", response_class=HTMLResponse)
async def findings_list(request: Request):
    qs = str(request.url.query)
    url = f"/auditor/findings?{qs}" if qs else "/auditor/findings"
    return RedirectResponse(url=url)


@router.post("/findings/clean")
async def clean_findings(
    request: Request,
    status: Optional[str] = None,
    severity: Optional[str] = None,
    category: Optional[str] = None,
    confidence: Optional[str] = None,
    file_path: Optional[str] = None,
    source: Optional[str] = None,
    system: Optional[str] = None,
):
    db = get_db(request)
    if db is None:
        return JSONResponse({"error": "No project configured"}, status_code=400)
    path_prefixes = None
    if system:
        sys_def = next((s for s in db.list_systems() if s["name"] == system), None)
        if sys_def:
            path_prefixes = sys_def["paths"]
    deleted = db.delete_findings_by_filter(
        status=status,
        severity=severity,
        category=category,
        confidence=confidence,
        file_path=file_path,
        source=source,
        path_prefixes=path_prefixes,
    )
    return JSONResponse({"deleted": deleted})


@router.get("/findings/export")
async def findings_export(
    request: Request,
    status: Optional[str] = None,
    severity: Optional[str] = None,
    category: Optional[str] = None,
    confidence: Optional[str] = None,
    file_path: Optional[str] = None,
    source: Optional[str] = None,
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    config = get_config(request)
    db = get_db(request)
    findings = db.list_findings(
        status=status,
        severity=severity,
        category=category,
        confidence=confidence,
        file_path=file_path,
        source=source,
    )
    scans = db.list_scans()
    stats = db.get_stats()

    wb = Workbook()

    # --- Sheet 1: Overview ---
    ws_overview = wb.active
    ws_overview.title = "Overview"
    bold = Font(bold=True)
    white_bold = Font(bold=True, color="FFFFFFFF")

    ws_overview.merge_cells("A1:C1")
    cell_title = ws_overview["A1"]
    cell_title.value = "Nytwatch Report"
    cell_title.font = bold

    project_name = Path(config.repo_path).name or config.repo_path
    ws_overview["A3"] = "Project"
    ws_overview["B3"] = project_name
    ws_overview["A4"] = "Generated"
    ws_overview["B4"] = datetime.now().strftime("%Y-%m-%d %H:%M")
    ws_overview["A5"] = "Confidence Threshold"
    ws_overview["B5"] = config.min_confidence
    ws_overview["A6"] = "File Extensions"
    ws_overview["B6"] = ", ".join(config.file_extensions)

    row = 8
    ws_overview.cell(row=row, column=1, value="Systems Configured").font = bold
    row += 1
    for system in db.list_systems():
        ws_overview.cell(row=row, column=2, value=system["name"])
        ws_overview.cell(row=row, column=3, value=", ".join(system["paths"]))
        row += 1

    row += 1
    ws_overview.cell(row=row, column=1, value="Severity Breakdown").font = bold
    row += 1
    severity_counts = stats.get("severity_counts", {})
    for sev in ["Critical", "High", "Medium", "Low", "Info"]:
        ws_overview.cell(row=row, column=2, value=sev)
        ws_overview.cell(row=row, column=3, value=severity_counts.get(sev.lower(), 0))
        row += 1

    row += 1
    ws_overview.cell(row=row, column=1, value="Scan History").font = bold
    row += 1
    for col_idx, header in enumerate(["System", "Files", "Findings", "Status", "Date"], start=2):
        cell = ws_overview.cell(row=row, column=col_idx, value=header)
        cell.font = bold
    row += 1
    for scan in scans:
        ws_overview.cell(row=row, column=2, value=scan.get("system_name", ""))
        ws_overview.cell(row=row, column=3, value=scan.get("files_scanned", 0))
        ws_overview.cell(row=row, column=4, value=scan.get("findings_count", 0))
        ws_overview.cell(row=row, column=5, value=scan.get("status", ""))
        ws_overview.cell(row=row, column=6, value=scan.get("started_at", ""))
        row += 1

    # --- Sheet 2: Findings ---
    ws_findings = wb.create_sheet("Findings")
    headers = [
        "Severity", "Source", "Title", "File", "Line Start", "Line End",
        "Category", "Confidence", "Status", "Description",
        "Suggested Fix", "Reasoning", "Test Description",
        "Created", "Reviewed",
    ]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_findings.cell(row=1, column=col_idx, value=header)
        cell.font = bold

    severity_fills = {
        "critical": PatternFill(start_color="FFDC2626", end_color="FFDC2626", fill_type="solid"),
        "high": PatternFill(start_color="FFEA580C", end_color="FFEA580C", fill_type="solid"),
        "medium": PatternFill(start_color="FFB45309", end_color="FFB45309", fill_type="solid"),
        "low": PatternFill(start_color="FF2563EB", end_color="FF2563EB", fill_type="solid"),
        "info": PatternFill(start_color="FF737686", end_color="FF737686", fill_type="solid"),
    }
    white_font = Font(color="FFFFFFFF")

    for row_idx, f in enumerate(findings, start=2):
        sev = f.get("severity", "")
        ws_findings.cell(row=row_idx, column=1, value=sev)
        ws_findings.cell(row=row_idx, column=2, value=f.get("source", "project"))
        ws_findings.cell(row=row_idx, column=3, value=f.get("title", ""))
        ws_findings.cell(row=row_idx, column=4, value=f.get("file_path", ""))
        ws_findings.cell(row=row_idx, column=5, value=f.get("line_start", 0))
        ws_findings.cell(row=row_idx, column=6, value=f.get("line_end", 0))
        ws_findings.cell(row=row_idx, column=7, value=f.get("category", ""))
        ws_findings.cell(row=row_idx, column=8, value=f.get("confidence", ""))
        ws_findings.cell(row=row_idx, column=9, value=f.get("status", ""))
        ws_findings.cell(row=row_idx, column=10, value=f.get("description", ""))
        ws_findings.cell(row=row_idx, column=11, value=f.get("suggested_fix", ""))
        ws_findings.cell(row=row_idx, column=12, value=f.get("reasoning", ""))
        ws_findings.cell(row=row_idx, column=13, value=f.get("test_description", ""))
        ws_findings.cell(row=row_idx, column=14, value=f.get("created_at", ""))
        ws_findings.cell(row=row_idx, column=15, value=f.get("reviewed_at", ""))

        sev_cell = ws_findings.cell(row=row_idx, column=1)
        if sev in severity_fills:
            sev_cell.fill = severity_fills[sev]
            sev_cell.font = white_font

    for col_idx in range(1, len(headers) + 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        for row_cells in ws_findings.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row_cells:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 2, 60)
        ws_findings.column_dimensions[column_letter].width = adjusted_width

    ws_findings.freeze_panes = "A2"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=nytwatch_findings.xlsx"},
    )


@router.get("/findings/{finding_id}", response_class=HTMLResponse)
async def finding_detail(request: Request, finding_id: str):
    return RedirectResponse(url=f"/auditor/findings/{finding_id}")


@router.post("/findings/{finding_id}/approve")
async def approve_finding(request: Request, finding_id: str):
    db = get_db(request)
    if db is None:
        return JSONResponse({"error": "Finding not found"}, status_code=404)
    finding = db.get_finding(finding_id)
    if not finding:
        return JSONResponse({"error": "Finding not found"}, status_code=404)
    if finding["status"] not in ("pending", "rejected"):
        return JSONResponse({"error": f"Cannot approve finding with status '{finding['status']}'"}, status_code=400)
    db.update_finding_status(finding_id, FindingStatus.APPROVED)
    logger.info("Finding %s approved", finding_id)
    return JSONResponse({"ok": True, "status": "approved"})


@router.post("/findings/{finding_id}/reject")
async def reject_finding(request: Request, finding_id: str):
    db = get_db(request)
    if db is None:
        return JSONResponse({"error": "Finding not found"}, status_code=404)
    finding = db.get_finding(finding_id)
    if not finding:
        return JSONResponse({"error": "Finding not found"}, status_code=404)
    if finding["status"] not in ("pending", "approved"):
        return JSONResponse({"error": f"Cannot reject finding with status '{finding['status']}'"}, status_code=400)
    db.update_finding_status(finding_id, FindingStatus.REJECTED)
    logger.info("Finding %s rejected", finding_id)
    return JSONResponse({"ok": True, "status": "rejected"})


@router.post("/findings/{finding_id}/toggle-test")
async def toggle_finding_test(request: Request, finding_id: str):
    db = get_db(request)
    if db is None:
        return JSONResponse({"error": "Finding not found"}, status_code=404)
    finding = db.get_finding(finding_id)
    if not finding:
        return JSONResponse({"error": "Finding not found"}, status_code=404)
    new_val = not bool(finding.get("include_test", 1))
    db.set_finding_include_test(finding_id, new_val)
    return JSONResponse({"ok": True, "include_test": new_val})


@router.post("/findings/{finding_id}/recheck")
async def recheck_finding(request: Request, finding_id: str):
    import asyncio
    db = get_db(request)
    if db is None:
        return JSONResponse({"error": "Not found"}, status_code=404)
    config = get_config(request)

    finding = db.get_finding(finding_id)
    if not finding:
        return JSONResponse({"error": "Not found"}, status_code=404)

    from nytwatch.analysis.engine import run_finding_recheck
    try:
        loop = asyncio.get_event_loop()
        still_valid, reason = await loop.run_in_executor(
            None, lambda: run_finding_recheck(finding, config.repo_path)
        )
    except Exception as e:
        logger.exception("Recheck error for %s", finding_id)
        return JSONResponse({"error": str(e)}, status_code=500)

    if not still_valid:
        db.update_finding_status(finding_id, FindingStatus.SUPERSEDED)
        logger.info("Finding %s recheck: no longer valid — superseded. Reason: %s", finding_id, reason)
    else:
        logger.info("Finding %s recheck: still valid. Reason: %s", finding_id, reason)

    return JSONResponse({"still_valid": still_valid, "reason": reason})


@router.get("/api/findings/{finding_id}/chat")
async def get_finding_chat(request: Request, finding_id: str):
    db = get_db(request)
    if db is None or not db.get_finding(finding_id):
        return JSONResponse({"error": "Not found"}, status_code=404)
    messages = db.get_finding_chat(finding_id)
    return JSONResponse({"messages": messages})


@router.post("/api/findings/{finding_id}/chat")
async def post_finding_chat(request: Request, finding_id: str):
    import asyncio
    db = get_db(request)
    if db is None:
        return JSONResponse({"error": "Not found"}, status_code=404)
    config = get_config(request)

    finding = db.get_finding(finding_id)
    if not finding:
        return JSONResponse({"error": "Not found"}, status_code=404)

    body = await request.json()
    user_message = (body.get("message") or "").strip()
    if not user_message:
        return JSONResponse({"error": "message is required"}, status_code=400)

    history = db.get_finding_chat(finding_id)

    from nytwatch.analysis.engine import run_finding_chat
    try:
        loop = asyncio.get_event_loop()
        display_text, updated_fields = await loop.run_in_executor(
            None,
            lambda: run_finding_chat(
                finding=finding,
                history=history,
                user_message=user_message,
                repo_path=config.repo_path,
            ),
        )
    except Exception as e:
        logger.exception("Finding chat error for %s", finding_id)
        return JSONResponse({"error": str(e)}, status_code=500)

    db.insert_chat_message(finding_id, "user", user_message)
    db.insert_chat_message(finding_id, "assistant", display_text)

    if updated_fields:
        db.update_finding_fields(finding_id, updated_fields)
        logger.info("Finding %s: chat updated fields %s", finding_id, list(updated_fields))

    return JSONResponse({"reply": display_text, "updated_fields": updated_fields})


# --- Scans ---

@router.get("/scans", response_class=HTMLResponse)
async def scans_list(request: Request):
    return RedirectResponse(url="/auditor/scans")


# --- System config API ---

@router.get("/api/browse-abs")
async def browse_absolute(path: str = "", file_ext: str = ""):
    """Browse the local filesystem by absolute path.

    When path is empty on Windows, returns the list of available drive letters.
    On Unix, falls through to root.

    If file_ext is provided (e.g. ".uproject"), matching files are included in
    the response alongside directories, with is_file=True.
    """
    import os
    import string

    _skip = {
        "$Recycle.Bin", "System Volume Information", "pagefile.sys",
        "hiberfil.sys", "swapfile.sys", "DumpStack.log.tmp",
        "Config.Msi", "MSOCache", "Recovery",
    }

    def _norm(p: Path) -> str:
        return str(p).replace("\\", "/")

    if not path:
        if os.name == "nt":
            drives = [
                f"{d}:/" for d in string.ascii_uppercase if Path(f"{d}:\\").exists()
            ]
            return JSONResponse({
                "path": "",
                "entries": [{"name": d.rstrip("/"), "path": d} for d in drives],
                "parent": None,
            })
        path = "/"

    target = Path(path).expanduser().resolve()
    if not target.exists() or not target.is_dir():
        return JSONResponse({"error": "Not a directory"}, status_code=404)

    entries = []
    try:
        for entry in sorted(target.iterdir(), key=lambda e: (e.is_file(), e.name.lower())):
            if entry.name in _skip:
                continue
            try:
                if entry.is_dir():
                    entries.append({"name": entry.name, "path": _norm(entry) + "/", "is_file": False})
                elif file_ext and entry.is_file() and entry.suffix.lower() == file_ext.lower():
                    entries.append({"name": entry.name, "path": _norm(entry), "is_file": True})
            except (PermissionError, OSError):
                pass
    except (PermissionError, OSError):
        pass

    # Parent logic: at a drive root (C:/) go back to the drives list ("")
    target_norm = _norm(target)
    parent_norm = _norm(target.parent)
    if os.name == "nt" and parent_norm == target_norm:
        parent = ""          # at drive root → up to drives list
    elif parent_norm == target_norm:
        parent = None        # at filesystem root
    else:
        parent = parent_norm + "/"

    return JSONResponse({
        "path": target_norm + "/",
        "entries": entries,
        "parent": parent,
    })


@router.get("/api/browse")
async def browse_directory(request: Request, path: str = "", base: str = ""):
    """Browse a directory tree.

    ``base`` is an optional absolute path to use as the root instead of the
    configured repo_path.  This is used by the setup wizard when the user is
    configuring a different repo than the currently active one.
    """
    _skip = {"Binaries", "Intermediate", "Saved", "DerivedDataCache", "__pycache__", "node_modules", ".git"}

    if base:
        repo = Path(base).expanduser().resolve()
        if not repo.exists() or not repo.is_dir():
            return JSONResponse({"error": "Base path not found or not a directory"}, status_code=400)
    else:
        config = get_config(request)
        repo = Path(config.repo_path).resolve()

    norm = path.replace("\\", "/").strip("/")
    target = (repo / norm).resolve() if norm else repo
    try:
        target.relative_to(repo)
    except ValueError:
        return JSONResponse({"error": "Invalid path"}, status_code=400)
    if not target.is_dir():
        return JSONResponse({"error": "Not a directory"}, status_code=404)

    entries = []
    try:
        for entry in sorted(target.iterdir(), key=lambda e: e.name.lower()):
            if entry.name.startswith(".") or entry.name in _skip or not entry.is_dir():
                continue
            rel = str(entry.relative_to(repo)).replace("\\", "/")
            entries.append({"name": entry.name, "path": rel + "/"})
    except PermissionError:
        pass

    current_rel = str(target.relative_to(repo)).replace("\\", "/") if target != repo else ""
    parent = None
    if current_rel:
        p = Path(current_rel).parent
        parent = "" if str(p) == "." else str(p).replace("\\", "/") + "/"

    return JSONResponse({"path": current_rel + "/" if current_rel else "", "entries": entries, "parent": parent})


@router.post("/api/open-folder")
async def open_folder(request: Request):
    """Open a local folder in the OS file explorer."""
    import subprocess
    import sys
    data = await request.json()
    path = data.get("path", "").strip()
    if not path:
        return JSONResponse({"error": "No path provided"}, status_code=400)
    p = Path(path)
    if not p.exists():
        return JSONResponse({"error": f"Path does not exist: {path}"}, status_code=404)
    try:
        if sys.platform == "win32":
            subprocess.Popen(["explorer", str(p)])
        elif sys.platform == "darwin":
            subprocess.Popen(["open", str(p)])
        else:
            subprocess.Popen(["xdg-open", str(p)])
        return JSONResponse({"ok": True})
    except Exception as exc:
        return JSONResponse({"error": str(exc)}, status_code=500)


@router.get("/api/systems")
async def get_systems_api(request: Request):
    db = get_db(request)
    return JSONResponse({"systems": db.list_systems() if db else []})


@router.get("/api/source-dirs")
async def get_source_dirs_api(request: Request):
    """Return active (non-ignored) source directories from the DB."""
    db = get_db(request)
    dirs = db.list_source_dirs() if db else []
    active = [d for d in dirs if d["source_type"] != "ignored"]
    return JSONResponse({"source_dirs": active})


@router.get("/api/source-dirs-all")
async def get_all_source_dirs_api(request: Request):
    """Return all source directories (active + ignored) from the DB."""
    db = get_db(request)
    return JSONResponse({"source_dirs": db.list_source_dirs() if db else []})


@router.post("/api/systems/append")
async def append_systems_api(request: Request):
    """Add new systems without touching existing ones."""
    db = get_db(request)
    if db is None:
        return JSONResponse({"error": "No project configured"}, status_code=400)
    body = await request.json()
    new_systems = body.get("systems", [])
    for s in new_systems:
        if not s.get("name", "").strip():
            return JSONResponse({"error": "System name cannot be empty"}, status_code=400)
        if not s.get("paths"):
            return JSONResponse({"error": f"System '{s['name']}' has no paths"}, status_code=400)

    existing = db.list_systems()
    existing_names = {s["name"] for s in existing}
    combined = list(existing) + [
        {
            "name": s["name"].strip(),
            "source_dir": s.get("source_dir") or "",
            "paths": s["paths"],
            "min_confidence": s.get("min_confidence") or None,
            "file_extensions": s.get("file_extensions") or None,
            "claude_fast_mode": s.get("claude_fast_mode"),
            "tracking_enabled": s.get("tracking_enabled", False),
            "tracking_verbosity": s.get("tracking_verbosity", "Standard"),
        }
        for s in new_systems
        if s["name"].strip() not in existing_names
    ]
    errs = _validate_systems(combined)
    if errs:
        return JSONResponse({"error": errs[0], "errors": errs}, status_code=400)
    db.replace_systems(combined)
    return JSONResponse({"ok": True})


@router.post("/api/systems")
async def save_systems_api(request: Request):
    db = get_db(request)
    if db is None:
        return JSONResponse({"error": "No project configured"}, status_code=400)
    body = await request.json()
    systems_data = body.get("systems", [])
    for s in systems_data:
        if not s.get("name", "").strip():
            return JSONResponse({"error": "System name cannot be empty"}, status_code=400)
        if not s.get("paths"):
            return JSONResponse({"error": f"System '{s['name']}' has no paths"}, status_code=400)
    # Build lookup of existing tracking config so it is preserved when not sent by caller
    existing_tracking = {
        s["name"]: {
            "tracking_enabled": s.get("tracking_enabled", False),
            "tracking_verbosity": s.get("tracking_verbosity", "Standard"),
        }
        for s in db.list_systems()
    }
    clean = [
        {
            "name": s["name"].strip(),
            "source_dir": s.get("source_dir") or "",
            "paths": s["paths"],
            "min_confidence": s.get("min_confidence") or None,
            "file_extensions": s.get("file_extensions") or None,
            "claude_fast_mode": s.get("claude_fast_mode"),
            "tracking_enabled": s.get(
                "tracking_enabled",
                existing_tracking.get(s["name"].strip(), {}).get("tracking_enabled", False),
            ),
            "tracking_verbosity": s.get(
                "tracking_verbosity",
                existing_tracking.get(s["name"].strip(), {}).get("tracking_verbosity", "Standard"),
            ),
        }
        for s in systems_data
    ]
    errs = _validate_systems(clean)
    if errs:
        return JSONResponse({"error": errs[0], "errors": errs}, status_code=400)
    db.replace_systems(clean)
    return JSONResponse({"ok": True})


# --- Project management API ---

@router.get("/api/projects")
async def list_projects(request: Request):
    from nytwatch.config import list_project_configs
    projects = list_project_configs()
    current_path = getattr(request.app.state, "config_path", "")
    return JSONResponse({"projects": projects, "current": current_path})


@router.post("/api/projects/switch")
async def switch_project(request: Request):
    from nytwatch.config import load_config, get_db_path
    body = await request.json()
    config_path_str = body.get("path", "").strip()
    if not config_path_str:
        return JSONResponse({"error": "No config path provided"}, status_code=400)
    p = Path(config_path_str)
    if not p.exists():
        return JSONResponse({"error": f"Config file not found: {config_path_str}"}, status_code=404)
    try:
        new_config = load_config(p)
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=400)

    # Switch in-memory config and database
    from nytwatch.database import Database
    old_db = request.app.state.db
    if old_db is not None:
        old_db.close()

    new_db = Database(get_db_path(new_config, p))
    new_db.init_schema()

    # Migrate systems from YAML if the new config carries them (legacy) and
    # the DB has none, or always restore when the YAML has authoritative data.
    if new_config.systems:
        new_db.replace_systems([
            {
                "name": s.name,
                "source_dir": s.source_dir,
                "paths": list(s.paths),
                "min_confidence": s.min_confidence,
                "file_extensions": list(s.file_extensions) if s.file_extensions else None,
                "claude_fast_mode": s.claude_fast_mode,
            }
            for s in new_config.systems
        ])
        logger.info(
            "Restored %d system(s) from YAML config for project: %s",
            len(new_config.systems), new_config.repo_path,
        )

    old_config = request.app.state.config
    old_repo = getattr(old_config, "repo_path", "") if old_config else ""

    request.app.state.config = new_config
    request.app.state.config_path = str(p)
    request.app.state.db = new_db
    from nytwatch.main import _read_tracking_active
    request.app.state.tracking_active = _read_tracking_active(new_config.repo_path) if new_config.repo_path else False

    # Update filesystem watcher to the new project
    watcher = getattr(request.app.state, "watcher", None)
    if watcher is not None:
        if old_repo:
            watcher.remove_watch(old_repo)
        if new_config.repo_path:
            watcher.add_watch(new_config.repo_path, new_db)

    from nytwatch.config import set_active_config_path
    set_active_config_path(p)

    logger.info("Switched to project: %s", new_config.repo_path)
    return JSONResponse({"ok": True, "repo_path": new_config.repo_path})


@router.delete("/api/projects")
async def delete_project(request: Request):
    """Delete the specified project config YAML and its SQLite database.

    If the deleted project is currently active, the server switches to the next
    available project (or a blank state) so it remains operational.
    """
    from nytwatch.config import (
        AuditorConfig, ACTIVE_POINTER_PATH, list_project_configs,
        set_active_config_path, get_db_path,
    )
    body = await request.json()
    config_path_str = body.get("path", "").strip()
    if not config_path_str:
        return JSONResponse({"error": "No config path provided"}, status_code=400)

    target = Path(config_path_str)
    if not target.exists():
        return JSONResponse({"error": f"Config file not found: {config_path_str}"}, status_code=404)

    current_path = getattr(request.app.state, "config_path", "")
    is_active = str(target) == current_path or target.resolve() == Path(current_path).resolve()

    # Determine the DB path before we delete the YAML
    try:
        from nytwatch.config import load_config
        target_config = load_config(target)
        target_db_path = get_db_path(target_config, target)
    except Exception:
        target_db_path = None

    # Close current DB if we're deleting the active project
    if is_active:
        try:
            request.app.state.db.close()
        except Exception:
            pass

    # Delete YAML config
    try:
        target.unlink()
    except Exception as e:
        return JSONResponse({"error": f"Failed to delete config file: {e}"}, status_code=500)

    # Delete associated SQLite database
    if target_db_path:
        db_path = Path(target_db_path)
        for suffix in ["", "-shm", "-wal"]:
            try:
                Path(str(db_path) + suffix).unlink(missing_ok=True)
            except Exception:
                pass

    if is_active:
        # Clear active pointer if it pointed to this file
        try:
            if ACTIVE_POINTER_PATH.exists():
                ACTIVE_POINTER_PATH.unlink()
        except Exception:
            pass

        # Switch to another available project, or start blank
        remaining = [p for p in list_project_configs() if p["path"] != config_path_str]
        if remaining:
            try:
                new_config = load_config(Path(remaining[0]["path"]))
                new_db_path = get_db_path(new_config, Path(remaining[0]["path"]))
                from nytwatch.database import Database
                new_db = Database(new_db_path)
                new_db.init_schema()
                request.app.state.config = new_config
                request.app.state.config_path = remaining[0]["path"]
                request.app.state.db = new_db
                set_active_config_path(Path(remaining[0]["path"]))
                logger.info("Deleted project, switched to: %s", new_config.repo_path)
                return JSONResponse({"ok": True, "switched_to": remaining[0]["path"]})
            except Exception as e:
                logger.warning("Could not switch to remaining project after delete: %s", e)

        # No remaining projects — go blank (no DB)
        request.app.state.config = AuditorConfig()
        request.app.state.config_path = ""
        request.app.state.db = None
        logger.info("Deleted last project, server is now unconfigured")
        return JSONResponse({"ok": True, "switched_to": None})

    logger.info("Deleted project config: %s", config_path_str)
    return JSONResponse({"ok": True})


def _validate_systems(systems: list[dict]) -> list[str]:
    """Return error strings for duplicate system names or duplicate paths across systems."""
    errors: list[str] = []
    seen_names: dict[str, bool] = {}
    seen_paths: dict[str, str] = {}
    for s in systems:
        name = (s.get("name") or "").strip()
        if not name:
            continue
        if name in seen_names:
            msg = f'System name: "{name}" conflicts with another system'
            if msg not in errors:
                errors.append(msg)
        else:
            seen_names[name] = True
        for p in s.get("paths", []):
            if not p:
                continue
            norm_p = p.replace("\\", "/")
            # Preserve file paths as-is; normalise directory paths to a trailing slash
            # so that "Source/Foo/" and "Source/Foo" are treated as the same directory.
            norm = norm_p if "." in norm_p.split("/")[-1] else norm_p.rstrip("/") + "/"
            if norm in seen_paths:
                owner = seen_paths[norm]
                if owner != name:
                    msg = f'Duplicate path: "{p}" is used by both "{owner}" and "{name}"'
                    if msg not in errors:
                        errors.append(msg)
            else:
                seen_paths[norm] = name
    return errors


def _make_build_config(build_data: dict):
    from nytwatch.config import BuildConfig
    from pathlib import Path as _Path
    import platform

    ue_dir = build_data.get("ue_installation_dir", "").strip()
    ue_cmd = build_data.get("ue_editor_cmd", "").strip()

    # Auto-derive ue_editor_cmd from the installation directory if not provided
    if ue_dir and not ue_cmd:
        exe = "UnrealEditor-Cmd.exe" if platform.system() == "Windows" else "UnrealEditor-Cmd"
        ue_cmd = str(_Path(ue_dir) / "Engine" / "Binaries" / "Win64" / exe)

    return BuildConfig(
        ue_installation_dir=ue_dir,
        ue_editor_cmd=ue_cmd,
        project_file=build_data.get("project_file", "").strip(),
        build_timeout_seconds=int(build_data.get("build_timeout_seconds", 1800)),
        test_timeout_seconds=int(build_data.get("test_timeout_seconds", 600)),
    )


@router.post("/api/projects/init")
async def init_project(request: Request):
    from nytwatch.config import (
        AuditorConfig, ScanSchedule, BuildConfig,
        NotificationConfig, save_full_config, DEFAULT_CONFIG_PATH, get_db_path,
    )
    import re
    body = await request.json()
    repo_path = body.get("repo_path", "").strip()
    if not repo_path:
        return JSONResponse({"error": "repo_path is required"}, status_code=400)
    if not Path(repo_path).expanduser().exists():
        return JSONResponse({"error": f"Path does not exist: {repo_path}"}, status_code=400)

    systems_data = body.get("systems", [])
    clean_systems = [
        {
            "name": s["name"].strip(),
            "source_dir": s.get("source_dir", ""),
            "paths": s["paths"],
            "min_confidence": s.get("min_confidence") or None,
            "file_extensions": s.get("file_extensions") or None,
            "claude_fast_mode": s.get("claude_fast_mode"),
        }
        for s in systems_data
        if s.get("name", "").strip()
    ]

    build_data = body.get("build", {})
    schedule_data = body.get("scan_schedule", {})

    config = AuditorConfig(
        repo_path=repo_path,
        build=_make_build_config(build_data),
        scan_schedule=ScanSchedule(
            incremental_interval_hours=int(schedule_data.get("incremental_interval_hours", 4)),
            rotation_enabled=bool(schedule_data.get("rotation_enabled", False)),
            rotation_interval_hours=int(schedule_data.get("rotation_interval_hours", 24)),
        ),
        claude_fast_mode=bool(body.get("claude_fast_mode", True)),
        min_confidence=body.get("min_confidence", "medium"),
    )

    config_path_str = body.get("config_path", "").strip()
    project_name = body.get("project_name", "").strip()
    if not config_path_str:
        # Derive from project_name if provided
        if project_name:
            slug = re.sub(r"[^a-z0-9_-]+", "-", project_name.lower()).strip("-")
            config_path_str = f"~/.nytwatch/{slug}.yaml"
    config_path = Path(config_path_str).expanduser() if config_path_str else DEFAULT_CONFIG_PATH

    config.project_name = project_name

    try:
        save_full_config(config, config_path)
    except Exception as e:
        return JSONResponse({"error": f"Failed to save config: {e}"}, status_code=500)

    # Ensure a DB exists for this project — create it now if this is the first project
    db = get_db(request)
    if db is None:
        from nytwatch.database import Database as _Database
        new_db_path = get_db_path(config, config_path)
        db = _Database(new_db_path)
        db.init_schema()
        request.app.state.db = db
        request.app.state.config = config
        request.app.state.config_path = str(config_path)

    if clean_systems:
        errs = _validate_systems(clean_systems)
        if errs:
            return JSONResponse({"error": errs[0], "errors": errs}, status_code=400)
        db.replace_systems(clean_systems)

    source_dirs = body.get("source_dirs", [])
    if source_dirs:
        valid_types = {"active", "ignored"}
        for entry in source_dirs:
            path = (entry.get("path") or "").strip()
            stype = (entry.get("source_type") or "active").strip()
            if stype not in valid_types:
                stype = "active"
            if path:
                db.upsert_source_dir(path, stype)

    from nytwatch.config import set_active_config_path
    set_active_config_path(config_path)

    logger.info("Project config created at: %s", config_path)
    return JSONResponse({"ok": True, "config_path": str(config_path).replace("\\", "/")})


@router.get("/api/git/branches")
async def git_branches_api(request: Request):
    """Return all local git branches for the configured repo, current branch first."""
    config = get_config(request)
    if not config.repo_path:
        return JSONResponse({"error": "No project configured"}, status_code=400)
    from nytwatch.pipeline.git_ops import get_local_branches
    branches = get_local_branches(config.repo_path)
    if not branches:
        return JSONResponse({"error": "Could not list git branches — is this a git repo?"}, status_code=400)
    # get_local_branches returns the current branch first, use that as fallback
    configured = config.git_branch or (branches[0] if branches else "")
    return JSONResponse({"branches": branches, "configured": configured})


@router.post("/api/config/branch")
async def set_branch_api(request: Request):
    """Change the configured git branch.

    Wipes all findings + batches + last_scan_commit so the new branch starts
    with a clean slate.  The server does NOT checkout the branch — the user
    controls their working tree.
    """
    from nytwatch.config import save_full_config
    from nytwatch.pipeline.git_ops import get_local_branches

    body = await request.json()
    branch = (body.get("branch") or "").strip()
    if not branch:
        return JSONResponse({"error": "branch is required"}, status_code=400)

    config = get_config(request)
    if not config.repo_path:
        return JSONResponse({"error": "No project configured"}, status_code=400)

    # Validate the branch actually exists locally
    branches = get_local_branches(config.repo_path)
    if branch not in branches:
        return JSONResponse(
            {"error": f"Branch '{branch}' not found in local repository"},
            status_code=400,
        )

    db = get_db(request)

    # Wipe findings, batches, and scan baseline — fresh start on the new branch
    wiped = db.wipe_findings()
    db.set_config("last_scan_commit", "")
    logger.info("Branch changed to '%s': wiped %d finding(s) and cleared scan baseline", branch, wiped)

    # Persist the new branch to YAML and update in-memory state
    from pydantic import BaseModel
    new_config = config.model_copy(update={"git_branch": branch})
    config_path_str = getattr(request.app.state, "config_path", "")
    if config_path_str:
        try:
            save_full_config(new_config, Path(config_path_str))
        except Exception as e:
            return JSONResponse({"error": f"Failed to save config: {e}"}, status_code=500)
    request.app.state.config = new_config

    return JSONResponse({"ok": True, "branch": branch, "wiped_findings": wiped})


@router.get("/api/detect-systems")
async def detect_systems_api(request: Request, repo_path: str = ""):
    from nytwatch.config import detect_systems_from_repo
    if not repo_path:
        config = get_config(request)
        repo_path = config.repo_path
    rp = Path(repo_path).expanduser()
    if not rp.exists():
        return JSONResponse({"error": f"Path does not exist: {repo_path}"}, status_code=400)
    candidates = detect_systems_from_repo(repo_path)
    return JSONResponse({"systems": candidates})


@router.get("/api/find-uproject")
async def find_uproject_api(request: Request, repo_path: str = ""):
    """Return the absolute path to the .uproject file in the given repo, if found."""
    if not repo_path:
        config = get_config(request)
        repo_path = config.repo_path
    rp = Path(repo_path).expanduser() if repo_path else None
    if not rp or not rp.exists():
        return JSONResponse({"error": f"Path does not exist: {repo_path}"}, status_code=400)
    matches = sorted(rp.glob("*.uproject"))
    if not matches:
        return JSONResponse({"uproject": None})
    return JSONResponse({"uproject": str(matches[0]).replace("\\", "/")})


@router.get("/api/validate-repo")
async def validate_repo_api(request: Request, repo_path: str = ""):
    """Lightweight repo validation — checks path exists and is a git repo. No filesystem scanning."""
    if not repo_path:
        return JSONResponse({"error": "repo_path is required"}, status_code=400)
    rp = Path(repo_path).expanduser()
    if not rp.exists():
        return JSONResponse({"error": f"Path does not exist: {repo_path}"}, status_code=400)
    if not rp.is_dir():
        return JSONResponse({"error": f"Not a directory: {repo_path}"}, status_code=400)
    if not (rp / ".git").exists():
        return JSONResponse({"error": f"Not a git repository (no .git folder): {repo_path}"}, status_code=400)
    return JSONResponse({"ok": True, "message": "Git repository found"})


@router.get("/api/detect-source-dirs")
async def detect_source_dirs_api(request: Request, repo_path: str = ""):
    """Return heuristically-classified source directories without touching the DB."""
    from nytwatch.scanner.source_detector import _heuristic_classify
    if not repo_path:
        config = get_config(request)
        repo_path = config.repo_path
    rp = Path(repo_path).expanduser() if repo_path else None
    if not rp or not rp.exists():
        return JSONResponse({"error": f"Path does not exist: {repo_path}"}, status_code=400)
    classified, unclassified = _heuristic_classify(rp)
    dirs = []
    for path, stype in sorted(classified.items()):
        dirs.append({"path": path, "source_type": stype, "auto": True})
    for path in sorted(unclassified):
        dirs.append({"path": path, "source_type": "project", "auto": False})
    return JSONResponse({"dirs": dirs})


def _build_suggest_systems_prompt(repo_path: str, active_dirs: list[dict]) -> str:
    dir_list = "\n".join(f"- {d['path']} (type: {d.get('source_type', 'project')})" for d in active_dirs)
    return f"""\
You are helping configure a code analysis tool for an Unreal Engine C++ project.

Repo root: {repo_path}

Active source directories to analyse:
{dir_list}

Your task: explore the directory structure under each active source directory using your tools (LS, Glob, etc.), then suggest logical "systems" — named groups of sub-paths that should be analysed together. Each system belongs to exactly one parent source directory.

Rules:
- Explore freely — go as deep as needed to find the right split point
- Feature folders (Combat, AI, Inventory, UI, etc.) at any depth → one system each
- A UE module wrapper folder (same name as the project/plugin) is not itself a split point — look inside it
- Small/leaf directories with no meaningful subdirs → one system at that path
- System names must be unique across ALL systems — no two systems may share the same name
- System names should be short and descriptive (use the feature folder name)
- "source_dir" must exactly match one of the active source directory paths listed above (with trailing slash)
- All paths must be DIRECTORIES (folders), never individual files — use forward slashes with a trailing slash
- Every path must be unique across ALL systems — no path may appear in more than one system's paths list
- UE module Public/Private rule: a UE module directory contains both a Public/ and a Private/ subfolder. There are two valid cases:
  1. If you want the entire module → use the module root path (e.g. "Module/") which covers all code inside it
  2. If you are splitting by feature folder inside the module → return BOTH sides explicitly, e.g. ["Module/Public/Combat/", "Module/Private/Combat/"]
  Never return a bare "Module/Public/" or "Module/Private/" without a feature subfolder — use the module root instead.

Return a JSON object — no markdown fences, no commentary:

{{
  "systems": [
    {{"name": "Combat",   "source_dir": "Source/Game/", "paths": ["Source/Game/ProjectName/Combat/"]}},
    {{"name": "AI",       "source_dir": "Source/Game/", "paths": ["Source/Game/ProjectName/AI/"]}},
    {{"name": "MyPlugin", "source_dir": "Plugins/MyPlugin/", "paths": ["Plugins/MyPlugin/Source/"]}}
  ]
}}\
"""


@router.post("/api/suggest-systems")
async def suggest_systems_api(request: Request):
    """Ask Claude to suggest logical scanning systems by exploring the repo with tools."""
    body = await request.json()
    repo_path = body.get("repo_path", "")
    source_dirs = body.get("source_dirs", [])  # [{path, source_type}]

    if not repo_path:
        config = get_config(request)
        repo_path = config.repo_path

    repo = Path(repo_path).expanduser()
    if not repo.exists():
        return JSONResponse({"error": f"Path does not exist: {repo_path}"}, status_code=400)

    active_dirs = [d for d in source_dirs if d.get("source_type") != "ignored"]
    if not active_dirs:
        return JSONResponse({"error": "No non-ignored source directories to analyse"}, status_code=400)

    prompt = _build_suggest_systems_prompt(str(repo), active_dirs)

    try:
        import subprocess as _sp
        from nytwatch.analysis.engine import call_claude, _extract_json
        try:
            # Agent mode: Claude explores the repo with its own tools
            raw = call_claude(prompt, fast=True, timeout=300, repo_path=str(repo), use_tools=True)
        except _sp.CalledProcessError as cpe:
            stderr_detail = (cpe.stderr or "").strip() or (cpe.stdout or "").strip()
            error_msg = stderr_detail or f"Claude CLI exited with code {cpe.returncode}"
            logger.error("suggest-systems: Claude CLI error (code %d): %s", cpe.returncode, error_msg)
            return JSONResponse({"error": f"Claude CLI error: {error_msg}"}, status_code=500)
        except FileNotFoundError:
            return JSONResponse({"error": "Claude CLI not found — ensure 'claude' is on PATH"}, status_code=500)
        data = _extract_json(raw)
        systems = data.get("systems", [])
        active_dir_paths = {d["path"] for d in active_dirs}
        # Longest-first so the most specific dir wins when inferring from a path
        active_dir_paths_sorted = sorted(active_dir_paths, key=len, reverse=True)
        valid = []
        for s in systems:
            name = (s.get("name") or "").strip()
            source_dir = (s.get("source_dir") or "").strip()
            paths = [p for p in (s.get("paths") or []) if isinstance(p, str) and p.strip()]
            if name and paths:
                # If Claude returned an unknown source_dir, infer from the first path
                if source_dir not in active_dir_paths:
                    first = paths[0].replace("\\", "/")
                    source_dir = next(
                        (d for d in active_dir_paths_sorted if first.startswith(d)), ""
                    )
                valid.append({"name": name, "source_dir": source_dir, "paths": paths})
        return JSONResponse({"systems": valid})
    except Exception as e:
        logger.exception("suggest-systems failed")
        return JSONResponse({"error": str(e)}, status_code=500)


def _build_suggest_paths_prompt(system_name: str, source_dir: str, repo_path: str) -> str:
    sd_root = source_dir.rstrip("/") + "/"
    return f"""\
You are configuring a code analysis tool for an Unreal Engine C++ project.

Repo root: {repo_path}
System name: "{system_name}"
Source directory: "{sd_root}"

Explore the directory structure under "{sd_root}" using your tools, then decide which paths this system should scan. Choose directories whose names and contents are related to the system name. When in doubt, prefer fewer, broader paths.

Rules:
- Only return paths that actually exist under "{sd_root}"
- If no subdirectory is clearly related, return the entire source dir: "{sd_root}"
- All paths must be DIRECTORIES (folders), never individual files — use forward slashes with a trailing slash
- Do not invent paths
- UE module Public/Private rule: a UE module directory contains both a Public/ and a Private/ subfolder. There are two valid cases:
  1. If this system covers the entire module → use the module root path (e.g. "Module/") which covers all code inside it
  2. If this system covers a specific feature inside the module → return BOTH sides explicitly, e.g. ["Module/Public/Combat/", "Module/Private/Combat/"]
  Never return a bare "Module/Public/" or "Module/Private/" without a feature subfolder — use the module root instead.

Return ONLY valid JSON (no fences, no commentary):
{{"paths": ["{sd_root}"]}}"""


@router.post("/api/suggest-paths")
async def suggest_paths_api(request: Request):
    """Ask Claude to suggest scan paths for a single named system inside its source_dir."""
    body = await request.json()
    system_name = (body.get("system_name") or "").strip()
    source_dir  = (body.get("source_dir")  or "").strip()

    if not system_name:
        return JSONResponse({"error": "system_name is required"}, status_code=400)
    if not source_dir:
        return JSONResponse({"error": "source_dir is required"}, status_code=400)

    config = get_config(request)
    repo_path = (body.get("repo_path") or config.repo_path or "").strip()
    if not repo_path:
        return JSONResponse({"error": "No repo path configured"}, status_code=400)

    repo = Path(repo_path).expanduser().resolve()
    sd_norm = source_dir.replace("\\", "/").rstrip("/")
    sd_path = (repo / sd_norm).resolve()
    sd_root = sd_norm + "/"

    if not sd_path.exists() or not sd_path.is_dir():
        return JSONResponse({"error": f"Source directory not found: {source_dir}"}, status_code=400)

    prompt = _build_suggest_paths_prompt(system_name, sd_root, str(repo))
    try:
        import subprocess as _sp
        from nytwatch.analysis.engine import call_claude, _extract_json
        try:
            # Agent mode: Claude explores the source dir with its own tools
            raw = call_claude(prompt, fast=True, timeout=300, repo_path=str(repo), use_tools=True)
        except _sp.CalledProcessError as cpe:
            stderr_detail = (cpe.stderr or "").strip() or (cpe.stdout or "").strip()
            return JSONResponse({"error": f"Claude CLI error: {stderr_detail or cpe.returncode}"}, status_code=500)
        except FileNotFoundError:
            return JSONResponse({"error": "Claude CLI not found — ensure 'claude' is on PATH"}, status_code=500)

        data = _extract_json(raw)
        raw_paths = [p for p in (data.get("paths") or []) if isinstance(p, str) and p.strip()]

        # Validate: every path must be under the source_dir
        valid = [p for p in raw_paths if p.replace("\\", "/").startswith(sd_root) or p.replace("\\", "/").rstrip("/") + "/" == sd_root]
        if not valid:
            valid = [sd_root]
        return JSONResponse({"paths": valid})
    except Exception:
        logger.exception("suggest-paths failed")
        return JSONResponse({"error": "Suggestion failed — check server logs"}, status_code=500)


@router.get("/api/config/status")
async def config_status(request: Request):
    from nytwatch.config import validate_config_errors, get_db_path, DEFAULT_CONFIG_PATH
    config = get_config(request)
    db = get_db(request)

    systems = db.list_systems() if db else []
    errors = validate_config_errors(config, systems=systems)
    config_path = getattr(request.app.state, "config_path", "")
    db_path = get_db_path(config, Path(config_path) if config_path else None) if db else None
    db_size = db_path.stat().st_size if db_path and db_path.exists() else 0
    last_commit = db.get_config("last_scan_commit", "") if db else ""

    repo = Path(config.repo_path).expanduser()
    config_path = getattr(request.app.state, "config_path", str(DEFAULT_CONFIG_PATH))
    return JSONResponse({
        "config_path": config_path,
        "repo_path": config.repo_path,
        "repo_exists": repo.exists(),
        "errors": errors,
        "last_commit": last_commit,
        "db_size_bytes": db_size,
    })


@router.post("/api/config/repair")
async def repair_config(request: Request):
    """Re-save the active config with all Pydantic defaults filled in."""
    from nytwatch.config import save_full_config
    config = get_config(request)
    config_path_raw = getattr(request.app.state, "config_path", "")
    if not config_path_raw:
        return JSONResponse(
            {"error": "No config file loaded — create a project first"},
            status_code=400,
        )
    config_path = Path(config_path_raw)
    try:
        save_full_config(config, config_path)
        logger.info("Config repaired at: %s", config_path)
        return JSONResponse({"ok": True})
    except Exception as e:
        logger.exception("Config repair failed")
        return JSONResponse({"error": str(e)}, status_code=500)


@router.post("/api/config/update")
async def update_config_api(request: Request):
    """Update build, schedule and quality settings in the active config file."""
    from nytwatch.config import save_full_config, ScanSchedule, AuditorConfig

    body = await request.json()
    config = get_config(request)
    config_path_str = getattr(request.app.state, "config_path", None)
    if not config_path_str:
        return JSONResponse(
            {"error": "No config file loaded — create a project first"},
            status_code=400,
        )

    build_data = body.get("build", {})
    sched_data = body.get("scan_schedule", {})

    new_config = AuditorConfig(
        repo_path=config.repo_path,
        data_dir=config.data_dir,
        notifications=config.notifications,
        build=_make_build_config({
            "ue_installation_dir": build_data.get("ue_installation_dir", config.build.ue_installation_dir),
            "project_file":        build_data.get("project_file",        config.build.project_file),
            "build_timeout_seconds": build_data.get("build_timeout_seconds", config.build.build_timeout_seconds),
            "test_timeout_seconds":  build_data.get("test_timeout_seconds",  config.build.test_timeout_seconds),
        }),
        scan_schedule=ScanSchedule(
            incremental_interval_hours=int(sched_data.get(
                "incremental_interval_hours", config.scan_schedule.incremental_interval_hours)),
            rotation_enabled=bool(sched_data.get(
                "rotation_enabled", config.scan_schedule.rotation_enabled)),
            rotation_interval_hours=int(sched_data.get(
                "rotation_interval_hours", config.scan_schedule.rotation_interval_hours)),
        ),
        claude_fast_mode=bool(body.get("claude_fast_mode", config.claude_fast_mode)),
        min_confidence=body.get("min_confidence", config.min_confidence),
        file_extensions=body.get("file_extensions", list(config.file_extensions)),
    )

    try:
        save_full_config(new_config, Path(config_path_str))
        request.app.state.config = new_config
        logger.info("Config updated at: %s", config_path_str)
        return JSONResponse({"ok": True})
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


@router.post("/scans/trigger")
async def trigger_scan(request: Request):
    config = get_config(request)
    db = get_db(request)
    if db is None:
        return JSONResponse({"error": "No project configured"}, status_code=400)
    db_path = db.db_path

    body = await request.json() if request.headers.get("content-type", "").startswith("application/json") else {}
    scan_type = body.get("scan_type", "full")

    # Accept system_id (preferred) or legacy system_name
    system_name: str | None = None
    system_id = body.get("system_id")
    if system_id is not None:
        matched = next((s for s in db.list_systems() if s["id"] == system_id), None)
        if matched is None:
            return JSONResponse({"error": f"System id {system_id!r} not found"}, status_code=404)
        system_name = matched["name"]
    else:
        system_name = body.get("system_name") or None

    # Reject if a scan is already running
    running = db.get_running_scan()
    if running:
        return JSONResponse(
            {"error": "A scan is already running", "scan_id": running["id"]},
            status_code=409,
        )

    from nytwatch.scan_state import canceller
    canceller.reset()

    def _run():
        from nytwatch.database import Database
        from nytwatch.scanner.scheduler import run_scan
        thread_db = Database(db_path)
        try:
            scan_id = run_scan(config, thread_db, scan_type=scan_type, system_name=system_name)
            logger.info("Scan completed: %s", scan_id)
        except Exception:
            logger.exception("Scan failed")
        finally:
            thread_db.close()

    thread = threading.Thread(target=_run, daemon=True)
    thread.start()
    return JSONResponse({"ok": True})


@router.delete("/scans/{scan_id}")
async def delete_scan(request: Request, scan_id: str):
    db = get_db(request)
    if db is None:
        return JSONResponse({"error": "Scan not found"}, status_code=404)
    scan = db.get_scan(scan_id)
    if not scan:
        return JSONResponse({"error": "Scan not found"}, status_code=404)
    if scan["status"] == "running":
        return JSONResponse({"error": "Cannot delete a running scan"}, status_code=400)
    db.delete_scan(scan_id)
    logger.info("Scan %s deleted", scan_id)
    return JSONResponse({"ok": True})


@router.post("/scans/cancel")
async def cancel_scan(request: Request):
    from nytwatch.scan_state import canceller
    db = get_db(request)
    if db is None:
        return JSONResponse({"error": "No project configured"}, status_code=400)

    if not canceller.is_cancelled:
        canceller.cancel()
        logger.info("Scan cancellation requested")

    from nytwatch.ws_manager import manager as ws_manager
    running = db.get_running_scan()
    ws_manager.push_scan_status(running=running is not None, scan=running, cancelling=True)

    if running:
        from nytwatch.models import ScanStatus, now_iso
        db.update_scan(running["id"], status=ScanStatus.CANCELLED, completed_at=now_iso())

    return JSONResponse({"ok": True})


# --- Settings ---

@router.get("/settings", response_class=HTMLResponse)
async def settings_page(request: Request):
    from nytwatch.config import DEFAULT_CONFIG_PATH
    db = get_db(request)
    config = get_config(request)
    source_dirs = db.list_source_dirs() if db else []
    active_dirs = [d for d in source_dirs if d["source_type"] != "ignored"]
    ignored_dirs = [d for d in source_dirs if d["source_type"] == "ignored"]

    # Count systems per source_dir for display; also build per-dir system list for reclassify
    all_systems = db.list_systems() if db else []
    sys_count: dict[str, int] = {}
    systems_by_dir: dict[str, list] = {}
    for s in all_systems:
        sd = (s.get("source_dir") or "").strip()
        sys_count[sd] = sys_count.get(sd, 0) + 1
        systems_by_dir.setdefault(sd, []).append({
            "name": s["name"],
            "source_dir": sd,
            "paths": s.get("paths") or [],
        })
    for d in active_dirs:
        d["system_count"] = sys_count.get(d["path"], 0)

    config_path = getattr(request.app.state, "config_path", "") or ""

    # Show whatever branch is stored in config — no subprocess on page load.
    configured_branch = config.git_branch or ""

    return templates.TemplateResponse(request, "settings.html", {
        "active_dirs": active_dirs,
        "ignored_dirs": ignored_dirs,
        "systems_by_dir": systems_by_dir,
        "config": config,
        "config_path": config_path,
        "configured_branch": configured_branch,
    })


@router.post("/settings/source-dirs")
async def update_source_dir(request: Request):
    db = get_db(request)
    if db is None:
        return JSONResponse({"error": "No project configured"}, status_code=400)
    body = await request.json()
    path = body.get("path", "").strip()
    source_type = body.get("source_type", "").strip()
    if not path or source_type not in ("active", "ignored"):
        return JSONResponse({"error": "Invalid path or source_type"}, status_code=400)
    db.upsert_source_dir(path, source_type)
    logger.info("Source dir updated: '%s' -> '%s'", path, source_type)
    return JSONResponse({"ok": True, "path": path, "source_type": source_type})


@router.delete("/settings/source-dirs")
async def delete_source_dir(request: Request):
    db = get_db(request)
    if db is None:
        return JSONResponse({"error": "No project configured"}, status_code=400)
    body = await request.json()
    path = body.get("path", "").strip()
    if not path:
        return JSONResponse({"error": "Invalid path"}, status_code=400)
    db.delete_source_dir(path)
    logger.info("Source dir deleted: '%s'", path)
    return JSONResponse({"ok": True, "path": path})


@router.post("/settings/source-dirs/bulk")
async def bulk_update_source_dirs(request: Request):
    """Apply a batch of upserts and deletes to source_dirs in one request.

    Body: {
        "upsert": [{"path": "...", "source_type": "active"|"ignored"}, ...],
        "delete": ["path1", "path2", ...]
    }
    """
    db = get_db(request)
    if db is None:
        return JSONResponse({"error": "No project configured"}, status_code=400)
    body = await request.json()
    valid_types = {"active", "ignored"}

    upserted = []
    for entry in body.get("upsert", []):
        path = (entry.get("path") or "").strip()
        stype = (entry.get("source_type") or "active").strip()
        if not path or stype not in valid_types:
            continue
        db.upsert_source_dir(path, stype)
        upserted.append(path)

    deleted = []
    for path in body.get("delete", []):
        path = (path or "").strip()
        if not path:
            continue
        db.delete_source_dir(path)
        deleted.append(path)

    logger.info("Bulk source-dirs: upserted %d, deleted %d", len(upserted), len(deleted))
    return JSONResponse({"ok": True, "upserted": upserted, "deleted": deleted})


@router.post("/settings/source-dirs/reclassify")
async def reclassify_source_dir(request: Request):
    """Replace systems for a source dir and purge all their findings."""
    db = get_db(request)
    if db is None:
        return JSONResponse({"error": "No project configured"}, status_code=400)
    body = await request.json()
    source_dir = (body.get("source_dir") or "").strip()
    old_system_names = [n for n in (body.get("old_system_names") or []) if n]
    new_systems = body.get("systems", [])

    if not source_dir:
        return JSONResponse({"error": "source_dir is required"}, status_code=400)

    for s in new_systems:
        if not (s.get("name") or "").strip():
            return JSONResponse({"error": "System name cannot be empty"}, status_code=400)
        if not [p for p in (s.get("paths") or []) if (p or "").strip()]:
            return JSONResponse({"error": f"System '{s.get('name', '')}' has no paths"}, status_code=400)

    # Delete findings (and their chats) for all old systems under this source dir
    deleted_findings = 0
    if old_system_names:
        placeholders = ",".join("?" * len(old_system_names))
        with db._lock:
            db.conn.execute(
                f"DELETE FROM finding_chats WHERE finding_id IN "
                f"(SELECT id FROM findings WHERE scan_id IN "
                f"(SELECT id FROM scans WHERE system_name IN ({placeholders})))",
                old_system_names,
            )
            cursor = db.conn.execute(
                f"DELETE FROM findings WHERE scan_id IN "
                f"(SELECT id FROM scans WHERE system_name IN ({placeholders}))",
                old_system_names,
            )
            db.conn.commit()
            deleted_findings = cursor.rowcount

    # Preserve tracking config for systems that survive the reclassify
    existing_tracking = {
        s["name"]: {
            "tracking_enabled": s.get("tracking_enabled", False),
            "tracking_verbosity": s.get("tracking_verbosity", "Standard"),
        }
        for s in db.list_systems()
    }

    other_systems = [s for s in db.list_systems() if (s.get("source_dir") or "").strip() != source_dir]
    new_clean = [
        {
            "name": s["name"].strip(),
            "source_dir": source_dir,
            "paths": [p for p in (s.get("paths") or []) if (p or "").strip()],
            "min_confidence": s.get("min_confidence") or None,
            "file_extensions": s.get("file_extensions") or None,
            "claude_fast_mode": s.get("claude_fast_mode"),
            "tracking_enabled": s.get(
                "tracking_enabled",
                existing_tracking.get(s["name"].strip(), {}).get("tracking_enabled", False),
            ),
            "tracking_verbosity": s.get(
                "tracking_verbosity",
                existing_tracking.get(s["name"].strip(), {}).get("tracking_verbosity", "Standard"),
            ),
        }
        for s in new_systems
    ]

    errs = _validate_systems(other_systems + new_clean)
    if errs:
        return JSONResponse({"error": errs[0], "errors": errs}, status_code=400)

    db.replace_systems(other_systems + new_clean)
    logger.info(
        "Reclassified source dir '%s': %d systems, %d findings deleted",
        source_dir, len(new_clean), deleted_findings,
    )
    return JSONResponse({"ok": True, "deleted_findings": deleted_findings, "systems_count": len(new_clean)})


# --- Batches ---

@router.get("/batches", response_class=HTMLResponse)
async def batches_list(request: Request):
    return RedirectResponse(url="/auditor/batches")


@router.get("/batches/{batch_id}", response_class=HTMLResponse)
async def batch_detail(request: Request, batch_id: str):
    return RedirectResponse(url=f"/auditor/batches/{batch_id}")


@router.post("/batch/apply")
async def apply_batch(request: Request):
    config = get_config(request)
    db = get_db(request)
    if db is None:
        return JSONResponse({"error": "No project configured"}, status_code=400)

    approved = db.get_approved_findings()
    if not approved:
        return JSONResponse({"error": "No approved findings to apply"}, status_code=400)

    from nytwatch.models import Batch, new_id
    batch = Batch(
        id=new_id(),
        finding_ids=[f["id"] for f in approved],
    )
    db.insert_batch(batch)

    for f in approved:
        db.set_finding_batch(f["id"], batch.id)

    db_path = db.db_path
    batch_id = batch.id

    def _run():
        from nytwatch.database import Database
        from nytwatch.pipeline.batch import run_batch_pipeline
        thread_db = Database(db_path)
        try:
            run_batch_pipeline(config, thread_db, batch_id)
        except Exception:
            logger.exception("Batch pipeline failed for %s", batch_id)
            thread_db.update_batch(batch_id, status=BatchStatus.FAILED)
        finally:
            thread_db.close()

    thread = threading.Thread(target=_run, daemon=True)
    thread.start()

    return JSONResponse({"ok": True, "batch_id": batch.id})


@router.post("/batches/{batch_id}/retry")
async def retry_batch(request: Request, batch_id: str):
    config = get_config(request)
    db = get_db(request)
    if db is None:
        return JSONResponse({"error": "No project configured"}, status_code=400)

    old_batch = db.get_batch(batch_id)
    if not old_batch:
        return JSONResponse({"error": "Batch not found"}, status_code=404)
    if old_batch["status"] != "failed":
        return JSONResponse({"error": "Only failed batches can be retried"}, status_code=400)

    # Reset findings back to approved so the new batch picks them up
    for fid in old_batch["finding_ids"]:
        db.update_finding_status(fid, FindingStatus.APPROVED)
        db.set_finding_batch(fid, None)

    from nytwatch.models import Batch, new_id
    new_batch = Batch(
        id=new_id(),
        finding_ids=old_batch["finding_ids"],
    )
    db.insert_batch(new_batch)
    for fid in old_batch["finding_ids"]:
        db.set_finding_batch(fid, new_batch.id)

    db_path = db.db_path
    new_batch_id = new_batch.id

    def _run():
        from nytwatch.database import Database
        from nytwatch.pipeline.batch import run_batch_pipeline
        thread_db = Database(db_path)
        try:
            run_batch_pipeline(config, thread_db, new_batch_id)
        except Exception:
            logger.exception("Batch pipeline failed for %s", new_batch_id)
            thread_db.update_batch(new_batch_id, status=BatchStatus.FAILED)
        finally:
            thread_db.close()

    thread = threading.Thread(target=_run, daemon=True)
    thread.start()

    logger.info("Retrying batch %s as new batch %s", batch_id, new_batch_id)
    return JSONResponse({"ok": True, "batch_id": new_batch_id})


# --- API ---

@router.get("/api/batches/{batch_id}/status")
async def api_batch_status(request: Request, batch_id: str):
    db = get_db(request)
    if db is None:
        return JSONResponse({"error": "No project configured"}, status_code=400)
    batch = db.get_batch(batch_id)
    if not batch:
        return JSONResponse({"error": "Batch not found"}, status_code=404)
    return JSONResponse({"status": batch["status"]})


@router.get("/api/stats")
async def api_stats(request: Request):
    db = get_db(request)
    if db is None:
        return JSONResponse({"findings": 0, "pending": 0, "approved": 0, "rejected": 0})
    return JSONResponse(db.get_stats())


_WS_PING_INTERVAL = 25  # seconds between server-side heartbeat pings


@router.websocket("/ws")
async def websocket_endpoint(websocket: WebSocket):
    import asyncio
    from nytwatch.ws_manager import manager as ws_manager
    from nytwatch.scan_state import canceller

    await ws_manager.connect(websocket)
    try:
        db = websocket.app.state.db
        running = db.get_running_scan() if db is not None else None
        await websocket.send_text(json.dumps({
            "type": "scan_status",
            "running": running is not None,
            "scan": running,
            "cancelling": canceller.is_cancelled,
        }))
        _ping = json.dumps({"type": "ping"})
        while True:
            try:
                await asyncio.wait_for(websocket.receive_text(), timeout=_WS_PING_INTERVAL)
            except asyncio.TimeoutError:
                # No message from client — send a keepalive ping so the
                # connection is not dropped by OS/browser idle timeouts.
                await websocket.send_text(_ping)
    except WebSocketDisconnect:
        pass
    except Exception:
        logger.exception("WebSocket error")
    finally:
        ws_manager.disconnect(websocket)


@router.websocket("/ws/tracking")
async def tracking_websocket(websocket: WebSocket, project_dir: str = ""):
    """WebSocket endpoint for the UE5 plugin to stream tracking events.

    The plugin connects here during PIE and sends:
      • ``session_open``  — once, on BeginPIE
      • ``event_batch``   — once per tick (only if changes occurred)
      • ``session_close`` — once, on EndPIE or crash

    ``project_dir`` (query param) identifies which project's DB and session
    directory to use.  The server writes this URL into NytwatchConfig.json so
    the plugin always connects with the right value.
    """
    from nytwatch.ws_manager import manager as ws_manager

    handler = getattr(websocket.app.state, "tracking_ws_handler", None)
    if handler is None:
        await websocket.close(code=1011)
        return

    db = websocket.app.state.db

    # If the project_dir doesn't match the active project, try to look it up
    # from the watcher's registered DBs so multi-project setups work.
    if project_dir:
        watcher = getattr(websocket.app.state, "watcher", None)
        if watcher is not None and project_dir in watcher._dbs:
            db = watcher._dbs[project_dir]

    await handler.handle(websocket, project_dir, ws_manager, db)


@router.get("/api/scans/{scan_id}/logs")
async def api_scan_logs(request: Request, scan_id: str, offset: int = 0):
    db = get_db(request)
    if db is None:
        return JSONResponse({"logs": [], "running": False, "total": 0})
    logs = db.get_scan_logs(scan_id, offset=offset)
    scan = db.get_scan(scan_id)
    running = scan["status"] == "running" if scan else False
    return JSONResponse({"logs": logs, "running": running, "total": len(logs)})


@router.get("/api/findings/stream")
async def api_findings_stream(request: Request, scan_id: str, offset: int = 0):
    """Return findings for a scan starting from offset N, respecting active filters."""
    db = get_db(request)
    findings = db.get_scan_findings_from(scan_id, offset)
    return JSONResponse({"findings": findings, "total": len(findings)})


@router.get("/api/scan-status")
async def api_scan_status(request: Request):
    from nytwatch.scan_state import canceller
    db = get_db(request)
    if db is None:
        return JSONResponse({"running": False, "scan": None, "cancelling": False})
    running = db.get_running_scan()
    return JSONResponse({
        "running": running is not None,
        "scan": running,
        "cancelling": canceller.is_cancelled,
    })


# ── Gameplay Tracker — PIE state ─────────────────────────────────────────────

@router.get("/api/nytwatch/pie-state")
async def api_pie_state(request: Request):
    config = get_config(request)
    watcher = getattr(request.app.state, "watcher", None)
    running = False
    if watcher is not None and config.repo_path:
        running = watcher.get_pie_state(config.repo_path)
    return JSONResponse({"running": running})


# ── Gameplay Tracker — Tracking start/stop ───────────────────────────────────

@router.post("/api/nytwatch/tracking/start")
async def api_tracking_start(request: Request):
    db = get_db(request)
    config = get_config(request)
    if db is None or not config.repo_path:
        return JSONResponse({"error": "No project configured"}, status_code=400)
    from nytwatch.tracking.config_writer import write_config
    request.app.state.tracking_active = True
    write_config(config.repo_path, db, tracking_active=True)
    return JSONResponse({"ok": True, "tracking_active": True})


@router.post("/api/nytwatch/tracking/stop")
async def api_tracking_stop(request: Request):
    db = get_db(request)
    config = get_config(request)
    if db is None or not config.repo_path:
        return JSONResponse({"error": "No project configured"}, status_code=400)
    from nytwatch.tracking.config_writer import write_config
    request.app.state.tracking_active = False
    write_config(config.repo_path, db, tracking_active=False)
    return JSONResponse({"ok": True, "tracking_active": False})


# ── Gameplay Tracker — Arm configuration ─────────────────────────────────────

@router.get("/api/nytwatch/arm")
async def api_get_arm_config(request: Request):
    db = get_db(request)
    if db is None:
        return JSONResponse({"error": "No project configured"}, status_code=400)
    systems = db.list_systems()
    tick_interval = float(db.get_config("tracking_tick_interval", "0.1"))
    scan_cap = int(db.get_config("tracking_scan_cap", "2000"))
    tracking_active = getattr(request.app.state, "tracking_active", False)
    return JSONResponse({
        "systems": systems,
        "tick_interval_seconds": tick_interval,
        "object_scan_cap": scan_cap,
        "tracking_active": tracking_active,
    })


@router.post("/api/nytwatch/arm")
async def api_post_arm_config(request: Request):
    db = get_db(request)
    config = get_config(request)
    if db is None or not config.repo_path:
        return JSONResponse({"error": "No project configured"}, status_code=400)
    body = await request.json()

    # Persist per-system tracking config
    for s in body.get("systems", []):
        name = s.get("name", "").strip()
        if not name:
            continue
        enabled = bool(s.get("tracking_enabled", False))
        verbosity = s.get("tracking_verbosity", "Standard")
        db.set_system_tracking(name, enabled, verbosity)

    # Persist global tracking settings
    tick = body.get("tick_interval_seconds")
    cap = body.get("object_scan_cap")
    if tick is not None:
        db.set_config("tracking_tick_interval", str(float(tick)))
    if cap is not None:
        db.set_config("tracking_scan_cap", str(int(cap)))

    from nytwatch.tracking.config_writer import write_config
    tracking_active = getattr(request.app.state, "tracking_active", False)
    write_config(config.repo_path, db, tracking_active=tracking_active)

    return JSONResponse({"ok": True})


# ── Gameplay Tracker — Per-file verbosity ────────────────────────────────────

@router.get("/api/nytwatch/systems/{system_name}/files")
async def api_get_system_files(request: Request, system_name: str):
    db = get_db(request)
    config = get_config(request)
    if db is None or not config.repo_path:
        return JSONResponse({"error": "No project configured"}, status_code=400)

    system = next((s for s in db.list_systems() if s["name"] == system_name), None)
    if system is None:
        return JSONResponse({"error": "System not found"}, status_code=404)

    overrides = {
        o["file_path"]: o["verbosity"]
        for o in db.get_file_verbosity_overrides(system_name)
    }

    repo = Path(config.repo_path)
    files = []
    for path_entry in system.get("paths", []):
        abs_path = Path(path_entry) if Path(path_entry).is_absolute() else repo / path_entry
        if abs_path.is_dir():
            for h_file in sorted(abs_path.rglob("*.h")):
                rel = str(h_file.relative_to(repo)).replace("\\", "/")
                files.append({
                    "file_path": rel,
                    "verbosity_override": overrides.get(rel) or overrides.get(str(h_file)),
                })
        elif abs_path.is_file() and abs_path.suffix == ".h":
            rel = str(abs_path.relative_to(repo)).replace("\\", "/")
            files.append({
                "file_path": rel,
                "verbosity_override": overrides.get(rel) or overrides.get(str(abs_path)),
            })

    return JSONResponse({"files": files})


@router.post("/api/nytwatch/systems/{system_name}/files/verbosity")
async def api_set_file_verbosity(request: Request, system_name: str):
    db = get_db(request)
    config = get_config(request)
    if db is None or not config.repo_path:
        return JSONResponse({"error": "No project configured"}, status_code=400)

    system = next((s for s in db.list_systems() if s["name"] == system_name), None)
    if system is None:
        return JSONResponse({"error": "System not found"}, status_code=404)

    body = await request.json()
    overrides = body.get("overrides", [])
    valid_verbosities = {"Critical", "Standard", "Verbose", "Ignore"}
    for o in overrides:
        v = o.get("verbosity") or ""
        if v and v not in valid_verbosities:
            return JSONResponse(
                {"error": f"Invalid verbosity: {v}"},
                status_code=400,
            )

    # Merge: upsert entries with a verbosity, delete entries with empty verbosity.
    # Files not present in the request are left untouched so that system-verbosity
    # changes never silently wipe overrides that the drawer didn't render.
    for o in overrides:
        fp = o.get("file_path", "").strip()
        v  = (o.get("verbosity") or "").strip()
        if not fp:
            continue
        if v:
            db.set_file_verbosity_override(system_name, fp, v)
        else:
            db.delete_file_verbosity_override(system_name, fp)

    from nytwatch.tracking.config_writer import write_config
    tracking_active = getattr(request.app.state, "tracking_active", False)
    write_config(config.repo_path, db, tracking_active=tracking_active)

    return JSONResponse({"ok": True})


# ── Gameplay Tracker — Sessions ───────────────────────────────────────────────



def _plugin_installed(repo_path: str | None) -> bool:
    """Return True if NytwatchAgent plugin directory exists in the project."""
    if not repo_path:
        return False
    return (Path(repo_path) / "Plugins" / "NytwatchAgent").is_dir()


@router.get("/api/nytwatch/plugin-check")
async def api_plugin_check(request: Request):
    config = get_config(request)
    return JSONResponse({"installed": _plugin_installed(config.repo_path)})


@router.get("/tracker", response_class=HTMLResponse)
async def tracker_redirect(request: Request):
    highlight = request.query_params.get("highlight")
    if highlight:
        return RedirectResponse(url=f"/tracker/sessions?highlight={highlight}")
    return RedirectResponse(url="/tracker/sessions")


async def _tracker_response(request: Request, active_tab: str, highlight: Optional[str] = None):
    db = get_db(request)
    if db is None:
        return RedirectResponse(url="/settings?setup=1")
    config = get_config(request)
    sessions = db.list_sessions(project_dir=config.repo_path or None)
    tracking_active = getattr(request.app.state, "tracking_active", False)
    tracking_tick = db.get_config("tracking_tick_interval", "0.1") if db else "0.1"
    tracking_cap = db.get_config("tracking_scan_cap", "2000") if db else "2000"
    plugin_installed = _plugin_installed(config.repo_path)

    watcher = getattr(request.app.state, "watcher", None)
    pie_recording = bool(watcher and config.repo_path and watcher.get_pie_state(config.repo_path))

    # Build grouped structure for the systems table
    all_systems = db.list_systems() if db else []
    all_source_dirs = db.list_source_dirs() if db else []
    source_dir_map = {d["path"]: d["source_type"] for d in all_source_dirs}
    sys_by_dir: dict[str, list] = {}
    for s in all_systems:
        sd = s.get("source_dir") or ""
        sys_by_dir.setdefault(sd, []).append(s)
    sorted_dirs = sorted(
        sys_by_dir.keys(),
        key=lambda sd: (1 if source_dir_map.get(sd) == "ignored" else 0, sd.lower()),
    )
    grouped_tracking = [
        {
            "source_dir": sd,
            "ignored": source_dir_map.get(sd) == "ignored",
            "systems": sorted(sys_by_dir[sd], key=lambda s: s["name"].lower()),
        }
        for sd in sorted_dirs
    ]
    has_tracking_grouping = any(g["source_dir"] for g in grouped_tracking)

    pie_armed  = [s["name"] for s in all_systems if s.get("tracking_enabled")] if pie_recording else []
    any_armed  = any(s.get("tracking_enabled") for s in all_systems)

    return templates.TemplateResponse(request, "tracker.html", {
        "sessions": sessions,
        "highlight": highlight,
        "active_tab": active_tab,
        "config": config,
        "tracking_active": tracking_active,
        "grouped_tracking": grouped_tracking,
        "has_tracking_grouping": has_tracking_grouping,
        "tracking_tick": tracking_tick,
        "tracking_cap": tracking_cap,
        "plugin_installed": plugin_installed,
        "pie_recording": pie_recording,
        "pie_armed": pie_armed,
        "any_armed": any_armed,
    })


@router.get("/tracker/sessions", response_class=HTMLResponse)
async def tracker_sessions(request: Request):
    highlight = request.query_params.get("highlight")
    return await _tracker_response(request, active_tab="sessions", highlight=highlight)


@router.get("/tracker/systems", response_class=HTMLResponse)
async def tracker_systems(request: Request):
    return await _tracker_response(request, active_tab="systems")


@router.get("/api/sessions")
async def api_list_sessions(
    request: Request,
    bookmarked_only: bool = False,
    q: str = "",
    limit: int = 100,
):
    db = get_db(request)
    if db is None:
        return JSONResponse({"sessions": []})
    config = get_config(request)
    sessions = db.list_sessions(
        project_dir=config.repo_path or None,
        bookmarked_only=bookmarked_only,
        name_filter=q,
        limit=limit,
    )
    return JSONResponse({"sessions": sessions})


@router.get("/api/sessions/{session_id}")
async def api_get_session(request: Request, session_id: str):
    db = get_db(request)
    if db is None:
        return JSONResponse({"error": "No project configured"}, status_code=400)
    session = db.get_session(session_id)
    if session is None:
        return JSONResponse({"error": "Session not found"}, status_code=404)
    return JSONResponse({"session": session})


@router.get("/api/sessions/{session_id}/content")
async def api_session_content(request: Request, session_id: str):
    db = get_db(request)
    if db is None:
        return JSONResponse({"error": "No project configured"}, status_code=400)
    session = db.get_session(session_id)
    if session is None:
        return JSONResponse({"error": "Session not found"}, status_code=404)
    file_path = session.get("file_path", "")
    try:
        content = Path(file_path).read_text(encoding="utf-8", errors="replace")
    except OSError as e:
        return JSONResponse({"error": f"Cannot read file: {e}"}, status_code=404)
    from fastapi.responses import PlainTextResponse
    return PlainTextResponse(content)


@router.post("/api/sessions/{session_id}/rename")
async def api_rename_session(request: Request, session_id: str):
    db = get_db(request)
    if db is None:
        return JSONResponse({"error": "No project configured"}, status_code=400)
    session = db.get_session(session_id)
    if session is None:
        return JSONResponse({"error": "Session not found"}, status_code=404)
    body = await request.json()
    new_name = (body.get("display_name") or "").strip()
    if not new_name:
        return JSONResponse({"error": "display_name cannot be empty"}, status_code=400)
    from nytwatch.tracking.session_store import rename_session
    rename_session(session_id, new_name, db)
    return JSONResponse({"ok": True})


@router.post("/api/sessions/{session_id}/bookmark")
async def api_bookmark_session(request: Request, session_id: str):
    db = get_db(request)
    if db is None:
        return JSONResponse({"error": "No project configured"}, status_code=400)
    session = db.get_session(session_id)
    if session is None:
        return JSONResponse({"error": "Session not found"}, status_code=404)
    body = await request.json()
    bookmarked = bool(body.get("bookmarked", False))
    from nytwatch.tracking.session_store import bookmark_session
    bookmark_session(session_id, bookmarked, db)
    return JSONResponse({"ok": True, "bookmarked": bookmarked})


@router.delete("/api/sessions/{session_id}")
async def api_delete_session(request: Request, session_id: str):
    db = get_db(request)
    if db is None:
        return JSONResponse({"error": "No project configured"}, status_code=400)
    session = db.get_session(session_id)
    if session is None:
        return JSONResponse({"error": "Session not found"}, status_code=404)
    if session.get("bookmarked"):
        return JSONResponse(
            {"error": "Cannot delete a bookmarked session. Unbookmark it first."},
            status_code=400,
        )
    from nytwatch.tracking.session_store import delete_session as _delete_session
    _delete_session(session_id, db)
    watcher = getattr(request.app.state, "watcher", None)
    if watcher is not None:
        watcher._ws.push_session_deleted(session_id)
    return JSONResponse({"ok": True})


# ═══════════════════════════════════════════════════════════════════════════════
# Project Management
# ═══════════════════════════════════════════════════════════════════════════════

def _pm_studio_path(request: Request) -> Optional[Path]:
    """Locate the studio root (dir containing production/) for the active project."""
    config = get_config(request)
    if not config or not getattr(config, "repo_path", ""):
        return None
    from nytwatch.pm.parser import find_studio_path
    return find_studio_path(config.repo_path)


# ── Pages ─────────────────────────────────────────────────────────────────────

@router.get("/pm", response_class=HTMLResponse)
async def pm_root(request: Request):
    return RedirectResponse(url="/pm/board")


@router.get("/pm/board", response_class=HTMLResponse)
async def pm_board(request: Request, sprint: Optional[int] = None):
    from nytwatch.pm.parser import load_sprints, load_milestones, sprint_to_dict, milestone_to_dict
    studio = _pm_studio_path(request)

    if studio is None:
        return templates.TemplateResponse(request, "project_management.html", {
            "active_tab": "board",
            "studio_path": None,
            "sprints": [],
            "active_sprint": None,
            "milestones": [],
        })

    sprints = load_sprints(studio)
    milestones = load_milestones(studio)

    if sprint is not None:
        active = next((s for s in sprints if s.number == sprint), None)
    else:
        # Default: first (lowest-numbered) sprint that is not closed
        active = next((s for s in sprints if not s.closed), None)
        if active is None and sprints:
            active = sprints[-1]

    return templates.TemplateResponse(request, "project_management.html", {
        "active_tab": "board",
        "studio_path": str(studio),
        "sprints": [sprint_to_dict(s) for s in sprints],
        "active_sprint": sprint_to_dict(active) if active else None,
        "milestones": [milestone_to_dict(m) for m in milestones],
    })


@router.get("/pm/milestones", response_class=HTMLResponse)
async def pm_milestones_page(request: Request):
    from nytwatch.pm.parser import load_sprints, load_milestones, sprint_to_dict, milestone_to_dict
    studio = _pm_studio_path(request)

    if studio is None:
        return templates.TemplateResponse(request, "project_management.html", {
            "active_tab": "milestones",
            "studio_path": None,
            "sprints": [],
            "active_sprint": None,
            "milestones": [],
        })

    sprints = load_sprints(studio)
    milestones = load_milestones(studio)

    return templates.TemplateResponse(request, "project_management.html", {
        "active_tab": "milestones",
        "studio_path": str(studio),
        "sprints": [sprint_to_dict(s) for s in sprints],
        "active_sprint": None,
        "milestones": [milestone_to_dict(m) for m in milestones],
    })


# ── Sprint API ────────────────────────────────────────────────────────────────

@router.post("/api/pm/sprints")
async def api_pm_create_sprint(request: Request):
    studio = _pm_studio_path(request)
    if studio is None:
        return JSONResponse({"error": "No studio path found"}, status_code=400)
    body = await request.json()
    sprint_n   = int(body.get("sprint_n", 1))
    goal       = body.get("goal", "")
    start_date = body.get("start_date", "")
    end_date   = body.get("end_date", "")
    from nytwatch.pm.writer import create_sprint_file
    from nytwatch.pm.parser import load_sprints, sprint_to_dict
    create_sprint_file(studio, sprint_n, goal, start_date, end_date)
    sprints = load_sprints(studio)
    created = next((s for s in sprints if s.number == sprint_n), None)
    return JSONResponse({"ok": True, "sprint": sprint_to_dict(created) if created else None})


@router.patch("/api/pm/sprints/{sprint_n}")
async def api_pm_update_sprint(request: Request, sprint_n: int):
    studio = _pm_studio_path(request)
    if studio is None:
        return JSONResponse({"error": "No studio path found"}, status_code=400)
    body = await request.json()
    from nytwatch.pm.writer import update_sprint_metadata
    ok = update_sprint_metadata(
        studio, sprint_n,
        goal=body.get("goal", ""),
        start_date=body.get("start_date", ""),
        end_date=body.get("end_date", ""),
    )
    return JSONResponse({"ok": ok})


@router.delete("/api/pm/sprints/{sprint_n}")
async def api_pm_delete_sprint(request: Request, sprint_n: int):
    studio = _pm_studio_path(request)
    if studio is None:
        return JSONResponse({"error": "No studio path found"}, status_code=400)
    from nytwatch.pm.writer import delete_sprint_file
    ok = delete_sprint_file(studio, sprint_n)
    return JSONResponse({"ok": ok})


@router.post("/api/pm/sprints/{sprint_n}/close")
async def api_pm_close_sprint(request: Request, sprint_n: int):
    studio = _pm_studio_path(request)
    if studio is None:
        return JSONResponse({"error": "No studio path found"}, status_code=400)
    body = await request.json()
    from nytwatch.pm.writer import close_sprint_file
    ok = close_sprint_file(studio, sprint_n, sprint_data=body.get("sprint"))
    return JSONResponse({"ok": ok})


# ── Task API ──────────────────────────────────────────────────────────────────

@router.post("/api/pm/sprints/{sprint_n}/tasks")
async def api_pm_create_task(request: Request, sprint_n: int):
    studio = _pm_studio_path(request)
    if studio is None:
        return JSONResponse({"error": "No studio path found"}, status_code=400)
    body = await request.json()
    task = {
        "id":                   body.get("id", ""),
        "name":                 body.get("name", ""),
        "owner":                body.get("owner", ""),
        "estimate_days":        body.get("estimate_days", 0),
        "dependencies":         body.get("dependencies", ""),
        "acceptance_criteria":  body.get("acceptance_criteria", ""),
        "sub_tasks":            body.get("sub_tasks", []),
        "priority":             body.get("priority", "should-have"),
        "status":               body.get("status", "backlog"),
        "sprint":               sprint_n,
        "blocker":              "",
        "completed":            "",
        "file":                 "",
    }
    from nytwatch.pm.writer import add_task_to_sprint
    add_task_to_sprint(studio, sprint_n, task)
    return JSONResponse({"ok": True, "task": task})


@router.patch("/api/pm/tasks/{sprint_n}/{task_id:path}")
async def api_pm_update_task(request: Request, sprint_n: int, task_id: str):
    studio = _pm_studio_path(request)
    if studio is None:
        return JSONResponse({"error": "No studio path found"}, status_code=400)
    body = await request.json()
    from nytwatch.pm.writer import update_task_in_sprint, move_task_between_sprints

    move_to = body.get("move_to_sprint")

    if move_to is not None and int(move_to) != sprint_n:
        task = {
            "id":                  task_id,
            "name":                body.get("name", ""),
            "owner":               body.get("owner", ""),
            "estimate_days":       body.get("estimate_days", 0),
            "dependencies":        body.get("dependencies", ""),
            "acceptance_criteria": body.get("acceptance_criteria", ""),
            "sub_tasks":           body.get("sub_tasks", []),
            "priority":            body.get("priority", "should-have"),
            "status":              body.get("status", "backlog"),
            "sprint":              sprint_n,
        }
        move_task_between_sprints(studio, task, sprint_n, int(move_to))
        return JSONResponse({"ok": True, "moved_to": move_to})

    # Status-only update (e.g. drag between Kanban columns)
    # task_name is sent by client so the checklist writer can locate the line.
    if "status" in body and "name" not in body:
        task = {
            "id":     task_id,
            "name":   body.get("task_name", ""),
            "status": body["status"],
        }
        update_task_in_sprint(studio, sprint_n, task)
        return JSONResponse({"ok": True})

    # Full task update
    old_name = body.get("old_name", "")
    task = {
        "id":                  task_id,
        "name":                body.get("name", ""),
        "owner":               body.get("owner", ""),
        "estimate_days":       body.get("estimate_days", 0),
        "dependencies":        body.get("dependencies", ""),
        "acceptance_criteria": body.get("acceptance_criteria", ""),
        "sub_tasks":           body.get("sub_tasks", []),
        "priority":            body.get("priority", "should-have"),
        "status":              body.get("status", "backlog"),
        "sprint":              sprint_n,
        "blocker":             body.get("blocker", ""),
        "completed":           body.get("completed", ""),
        "file":                body.get("file", ""),
    }
    update_task_in_sprint(studio, sprint_n, task, old_name=old_name)
    return JSONResponse({"ok": True})


@router.post("/api/pm/subtask-status")
async def api_pm_subtask_status(request: Request):
    """Toggle a single sub-task marker without touching any other field.
    Uses a flat POST URL to avoid conflict with the {task_id:path} task route."""
    studio = _pm_studio_path(request)
    if studio is None:
        return JSONResponse({"error": "No studio path found"}, status_code=400)
    body = await request.json()
    from nytwatch.pm.writer import set_subtask_done_in_sprint
    ok = set_subtask_done_in_sprint(
        studio,
        int(body.get("sprint_n", 0)),
        body.get("task_id", ""),
        body.get("task_name", ""),
        body.get("subtask_id", ""),
        bool(body.get("done", False)),
    )
    return JSONResponse({"ok": ok})


@router.delete("/api/pm/tasks/{sprint_n}/{task_id:path}")
async def api_pm_delete_task(request: Request, sprint_n: int, task_id: str):
    studio = _pm_studio_path(request)
    if studio is None:
        return JSONResponse({"error": "No studio path found"}, status_code=400)
    from nytwatch.pm.writer import remove_task_from_sprint
    remove_task_from_sprint(studio, sprint_n, task_id)
    return JSONResponse({"ok": True})


# ── Milestone API ─────────────────────────────────────────────────────────────

@router.post("/api/pm/milestones")
async def api_pm_create_milestone(request: Request):
    studio = _pm_studio_path(request)
    if studio is None:
        return JSONResponse({"error": "No studio path found"}, status_code=400)
    body = await request.json()
    slug        = re.sub(r'[^a-z0-9-]', '-', body.get("name", "milestone").lower()).strip('-')
    name        = body.get("name", "")
    target_date = body.get("target_date", "")
    goal        = body.get("goal", "")
    sprints     = [int(n) for n in body.get("sprints", [])]
    status      = body.get("status", "Planned")
    from nytwatch.pm.writer import create_milestone
    from nytwatch.pm.parser import load_milestones, milestone_to_dict
    create_milestone(studio, slug, name, target_date, goal, sprints, status)
    ms = load_milestones(studio)
    created = next((m for m in ms if m.slug == slug), None)
    return JSONResponse({"ok": True, "milestone": milestone_to_dict(created) if created else None})


@router.patch("/api/pm/milestones/{slug}")
async def api_pm_update_milestone(request: Request, slug: str):
    studio = _pm_studio_path(request)
    if studio is None:
        return JSONResponse({"error": "No studio path found"}, status_code=400)
    body = await request.json()
    from nytwatch.pm.writer import update_milestone
    ok = update_milestone(
        studio, slug,
        name=body.get("name", ""),
        target_date=body.get("target_date", ""),
        goal=body.get("goal", ""),
        sprints=[int(n) for n in body.get("sprints", [])],
        status=body.get("status", "Planned"),
    )
    return JSONResponse({"ok": ok})


@router.delete("/api/pm/milestones/{slug}")
async def api_pm_delete_milestone(request: Request, slug: str):
    studio = _pm_studio_path(request)
    if studio is None:
        return JSONResponse({"error": "No studio path found"}, status_code=400)
    from nytwatch.pm.writer import delete_milestone
    ok = delete_milestone(studio, slug)
    return JSONResponse({"ok": ok})


# ═══════════════════════════════════════════════════════════════════════════════
# Wiki
# ═══════════════════════════════════════════════════════════════════════════════

def _wiki_path(request: Request) -> Optional[Path]:
    """Locate the wiki directory for the active project.

    Tries (in order):
      1. <studio>/production/wiki/   — consistent with the existing PM layout
      2. <repo_root>/planning/wiki/  — simple flat layout for new projects
    """
    config = get_config(request)
    if not config or not getattr(config, "repo_path", ""):
        return None

    studio = _pm_studio_path(request)
    if studio is not None:
        p = studio / "production" / "wiki"
        if p.exists():
            return p

    repo = Path(config.repo_path)
    p2 = repo / "planning" / "wiki"
    if p2.exists():
        return p2

    return None


# ═══════════════════════════════════════════════════════════════════════════════
# Docs
# ═══════════════════════════════════════════════════════════════════════════════

def _design_path(request: Request) -> Optional[str]:
    """Return the repo_path for the active project, used by docs_parser."""
    config = get_config(request)
    if not config or not getattr(config, "repo_path", ""):
        return None
    return config.repo_path


@router.get("/docs", response_class=HTMLResponse)
async def docs_page(request: Request, doc: Optional[str] = None):
    from nytwatch.pm.docs_parser import load_design_docs, doc_to_dict, _planning_root

    repo_path = _design_path(request)

    if repo_path is None:
        return templates.TemplateResponse(request, "docs.html", {
            "design_path": None,
            "docs": [],
            "selected_doc": None,
        })

    docs = load_design_docs(repo_path)
    planning = _planning_root(repo_path)

    selected = None
    if doc:
        selected = next((d for d in docs if d.slug == doc), None)

    return templates.TemplateResponse(request, "docs.html", {
        "design_path": str(planning) if planning else None,
        "docs": [doc_to_dict(d) for d in docs],
        "selected_doc": doc_to_dict(selected) if selected else None,
    })


# ═══════════════════════════════════════════════════════════════════════════════
# Wiki
# ═══════════════════════════════════════════════════════════════════════════════

@router.get("/wiki", response_class=HTMLResponse)
async def wiki_page(request: Request, doc: Optional[str] = None):
    from nytwatch.pm.wiki_parser import load_wiki_docs, doc_to_dict

    wiki_path = _wiki_path(request)

    if wiki_path is None:
        return templates.TemplateResponse(request, "wiki.html", {
            "wiki_path": None,
            "docs": [],
            "selected_doc": None,
        })

    docs = load_wiki_docs(wiki_path)

    selected = None
    if doc:
        selected = next((d for d in docs if d.slug == doc), None)

    return templates.TemplateResponse(request, "wiki.html", {
        "wiki_path": str(wiki_path),
        "docs": [doc_to_dict(d) for d in docs],
        "selected_doc": doc_to_dict(selected) if selected else None,
    })
