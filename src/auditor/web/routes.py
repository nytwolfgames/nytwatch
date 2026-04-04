from __future__ import annotations

import json
import logging
import threading
from io import BytesIO
from datetime import datetime
from pathlib import Path
from typing import Optional

from fastapi import APIRouter, Request, WebSocket, WebSocketDisconnect
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse, StreamingResponse
from fastapi.templating import Jinja2Templates

from auditor.database import Database
from auditor.models import BatchStatus, FindingStatus, now_iso

logger = logging.getLogger(__name__)

router = APIRouter()

templates = Jinja2Templates(
    directory=str(Path(__file__).parent / "templates")
)

# ── Template globals ─────────────────────────────────────────────────────────
# Callables receive `request` so they always reflect the current active project
# without requiring every route to explicitly pass these values.

def _active_project_name(request: "Request") -> str:
    """Return the short project name (repo folder) for the active project."""
    config = getattr(request.app.state, "config", None)
    if config and getattr(config, "repo_path", ""):
        return Path(config.repo_path).name
    return ""

def _active_config_path(request: "Request") -> str:
    return getattr(request.app.state, "config_path", "") or ""

def _active_repo_path(request: "Request") -> str:
    config = getattr(request.app.state, "config", None)
    return getattr(config, "repo_path", "") if config else ""

templates.env.globals["active_project_name"] = _active_project_name
templates.env.globals["active_config_path"] = _active_config_path
templates.env.globals["active_repo_path"] = _active_repo_path


def get_db(request: Request) -> Database:
    return request.app.state.db


def get_config(request: Request):
    return request.app.state.config


# --- Dashboard ---

def _infer_source_dirs(systems: list[dict], source_dir_paths: set[str]) -> list[dict]:
    """For systems with an empty source_dir, infer it by longest-prefix match against
    the known source directories.  Returns the full list (modified in place for matches)."""
    # Normalise every source dir to "path/with/trailing/slash"
    sd_normed = {sd: sd.replace("\\", "/").rstrip("/") + "/" for sd in source_dir_paths}

    result = []
    for s in systems:
        if s.get("source_dir"):
            result.append(s)
            continue
        best_sd, best_len = "", 0
        for path in s.get("paths", []):
            p = path.replace("\\", "/")
            if not p.endswith("/"):
                p += "/"
            for sd, sd_norm in sd_normed.items():
                if p.startswith(sd_norm) and len(sd_norm) > best_len:
                    best_sd, best_len = sd, len(sd_norm)
        result.append(dict(s, source_dir=best_sd) if best_sd else s)
    return result


@router.get("/", response_class=HTMLResponse)
async def dashboard(request: Request):
    db = get_db(request)
    config = get_config(request)

    # First-run: no project configured yet → redirect to setup wizard.
    # Only redirect when there is genuinely no active project (no repo_path).
    # A configured project with no systems yet is valid and should show the dashboard.
    if not config.repo_path:
        return RedirectResponse(url="/settings?setup=1")

    db_systems = db.list_systems()
    all_source_dirs = db.list_source_dirs()
    source_dir_map = {d["path"]: d["source_type"] for d in all_source_dirs}

    # ── One-time repair: assign source_dir to systems that are missing it ──────
    needs_repair = any(not s.get("source_dir") for s in db_systems)
    if needs_repair and source_dir_map:
        repaired = _infer_source_dirs(db_systems, set(source_dir_map.keys()))
        db.replace_systems([
            {
                "name": s["name"],
                "source_dir": s.get("source_dir") or "",
                "paths": s.get("paths", []),
                "min_confidence": s.get("min_confidence"),
                "file_extensions": s.get("file_extensions"),
                "claude_fast_mode": s.get("claude_fast_mode"),
            }
            for s in repaired
        ])
        db_systems = db.list_systems()
        logger.info("Repaired source_dir for %d system(s)",
                    sum(1 for s in repaired if s.get("source_dir")))

    # ── Build grouped structure ───────────────────────────────────────────────
    sys_by_dir: dict[str, list] = {}
    for s in db_systems:
        sd = s.get("source_dir") or ""
        sys_by_dir.setdefault(sd, []).append({
            "id": s["id"],
            "name": s["name"],
            "source_dir": sd,
            "count": db.count_findings_for_path_prefixes(s["paths"]),
        })

    # Sort groups: active before ignored, then alphabetically by source_dir name
    all_source_dirs = list(sys_by_dir.keys())
    all_source_dirs.sort(key=lambda sd: (
        1 if source_dir_map.get(sd) == "ignored" else 0,  # active first
        sd.lower(),
    ))

    grouped_systems = []
    for sd in all_source_dirs:
        systems_in_group = sorted(sys_by_dir[sd], key=lambda s: s["name"].lower())
        grouped_systems.append({
            "source_dir": sd,
            "ignored": source_dir_map.get(sd) == "ignored",
            "systems": systems_in_group,
        })

    has_grouping = any(g["source_dir"] for g in grouped_systems)

    # Flat list in the same order — used as JS index baseline
    systems = [s for g in grouped_systems for s in g["systems"]]

    stats = db.get_stats()
    batches = db.list_batches(limit=5)
    return templates.TemplateResponse(request, "dashboard.html", {
        "stats": stats,
        "batches": batches,
        "systems": systems,
        "grouped_systems": grouped_systems,
        "has_grouping": has_grouping,
    })


# --- Findings ---

@router.get("/findings", response_class=HTMLResponse)
async def findings_list(
    request: Request,
    status: Optional[str] = None,
    severity: Optional[str] = None,
    category: Optional[str] = None,
    confidence: Optional[str] = None,
    file_path: Optional[str] = None,
    source: Optional[str] = None,
    system: Optional[str] = None,
):
    config = get_config(request)
    db = get_db(request)

    path_prefixes = None
    if system:
        sys_def = next((s for s in db.list_systems() if s["name"] == system), None)
        if sys_def:
            path_prefixes = sys_def["paths"]

    findings = db.list_findings(
        status=status,
        severity=severity,
        category=category,
        confidence=confidence,
        file_path=file_path,
        source=source,
        path_prefixes=path_prefixes,
    )
    approved_count = len(db.get_approved_findings())
    filters = {
        "status": status,
        "severity": severity,
        "category": category,
        "confidence": confidence,
        "file_path": file_path,
        "source": source,
        "system": system,
    }
    return templates.TemplateResponse(request, "findings_list.html", {
        "findings": findings,
        "filters": filters,
        "approved_count": approved_count,
        "systems": db.list_systems(),  # [{name, source_dir, paths, ...}]
    })


@router.get("/findings/export")
async def findings_export(
    request: Request,
    status: Optional[str] = None,
    severity: Optional[str] = None,
    category: Optional[str] = None,
    confidence: Optional[str] = None,
    file_path: Optional[str] = None,
    source: Optional[str] = None,
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    config = get_config(request)
    db = get_db(request)
    findings = db.list_findings(
        status=status,
        severity=severity,
        category=category,
        confidence=confidence,
        file_path=file_path,
        source=source,
    )
    scans = db.list_scans()
    stats = db.get_stats()

    wb = Workbook()

    # --- Sheet 1: Overview ---
    ws_overview = wb.active
    ws_overview.title = "Overview"
    bold = Font(bold=True)
    white_bold = Font(bold=True, color="FFFFFFFF")

    ws_overview.merge_cells("A1:C1")
    cell_title = ws_overview["A1"]
    cell_title.value = "Code Auditor Report"
    cell_title.font = bold

    project_name = Path(config.repo_path).name or config.repo_path
    ws_overview["A3"] = "Project"
    ws_overview["B3"] = project_name
    ws_overview["A4"] = "Generated"
    ws_overview["B4"] = datetime.now().strftime("%Y-%m-%d %H:%M")
    ws_overview["A5"] = "Confidence Threshold"
    ws_overview["B5"] = config.min_confidence
    ws_overview["A6"] = "File Extensions"
    ws_overview["B6"] = ", ".join(config.file_extensions)

    row = 8
    ws_overview.cell(row=row, column=1, value="Systems Configured").font = bold
    row += 1
    for system in db.list_systems():
        ws_overview.cell(row=row, column=2, value=system["name"])
        ws_overview.cell(row=row, column=3, value=", ".join(system["paths"]))
        row += 1

    row += 1
    ws_overview.cell(row=row, column=1, value="Severity Breakdown").font = bold
    row += 1
    severity_counts = stats.get("severity_counts", {})
    for sev in ["Critical", "High", "Medium", "Low", "Info"]:
        ws_overview.cell(row=row, column=2, value=sev)
        ws_overview.cell(row=row, column=3, value=severity_counts.get(sev.lower(), 0))
        row += 1

    row += 1
    ws_overview.cell(row=row, column=1, value="Scan History").font = bold
    row += 1
    for col_idx, header in enumerate(["System", "Files", "Findings", "Status", "Date"], start=2):
        cell = ws_overview.cell(row=row, column=col_idx, value=header)
        cell.font = bold
    row += 1
    for scan in scans:
        ws_overview.cell(row=row, column=2, value=scan.get("system_name", ""))
        ws_overview.cell(row=row, column=3, value=scan.get("files_scanned", 0))
        ws_overview.cell(row=row, column=4, value=scan.get("findings_count", 0))
        ws_overview.cell(row=row, column=5, value=scan.get("status", ""))
        ws_overview.cell(row=row, column=6, value=scan.get("started_at", ""))
        row += 1

    # --- Sheet 2: Findings ---
    ws_findings = wb.create_sheet("Findings")
    headers = [
        "Severity", "Source", "Title", "File", "Line Start", "Line End",
        "Category", "Confidence", "Status", "Description",
        "Suggested Fix", "Reasoning", "Test Description",
        "Created", "Reviewed",
    ]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_findings.cell(row=1, column=col_idx, value=header)
        cell.font = bold

    severity_fills = {
        "critical": PatternFill(start_color="FFDC2626", end_color="FFDC2626", fill_type="solid"),
        "high": PatternFill(start_color="FFEA580C", end_color="FFEA580C", fill_type="solid"),
        "medium": PatternFill(start_color="FFB45309", end_color="FFB45309", fill_type="solid"),
        "low": PatternFill(start_color="FF2563EB", end_color="FF2563EB", fill_type="solid"),
        "info": PatternFill(start_color="FF737686", end_color="FF737686", fill_type="solid"),
    }
    white_font = Font(color="FFFFFFFF")

    for row_idx, f in enumerate(findings, start=2):
        sev = f.get("severity", "")
        ws_findings.cell(row=row_idx, column=1, value=sev)
        ws_findings.cell(row=row_idx, column=2, value=f.get("source", "project"))
        ws_findings.cell(row=row_idx, column=3, value=f.get("title", ""))
        ws_findings.cell(row=row_idx, column=4, value=f.get("file_path", ""))
        ws_findings.cell(row=row_idx, column=5, value=f.get("line_start", 0))
        ws_findings.cell(row=row_idx, column=6, value=f.get("line_end", 0))
        ws_findings.cell(row=row_idx, column=7, value=f.get("category", ""))
        ws_findings.cell(row=row_idx, column=8, value=f.get("confidence", ""))
        ws_findings.cell(row=row_idx, column=9, value=f.get("status", ""))
        ws_findings.cell(row=row_idx, column=10, value=f.get("description", ""))
        ws_findings.cell(row=row_idx, column=11, value=f.get("suggested_fix", ""))
        ws_findings.cell(row=row_idx, column=12, value=f.get("reasoning", ""))
        ws_findings.cell(row=row_idx, column=13, value=f.get("test_description", ""))
        ws_findings.cell(row=row_idx, column=14, value=f.get("created_at", ""))
        ws_findings.cell(row=row_idx, column=15, value=f.get("reviewed_at", ""))

        sev_cell = ws_findings.cell(row=row_idx, column=1)
        if sev in severity_fills:
            sev_cell.fill = severity_fills[sev]
            sev_cell.font = white_font

    for col_idx in range(1, len(headers) + 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        for row_cells in ws_findings.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row_cells:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 2, 60)
        ws_findings.column_dimensions[column_letter].width = adjusted_width

    ws_findings.freeze_panes = "A2"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=code_auditor_findings.xlsx"},
    )


@router.get("/findings/{finding_id}", response_class=HTMLResponse)
async def finding_detail(request: Request, finding_id: str):
    db = get_db(request)
    finding = db.get_finding(finding_id)
    if not finding:
        return HTMLResponse("<h1>Finding not found</h1>", status_code=404)
    return templates.TemplateResponse(request, "finding_detail.html", {
        "finding": finding,
    })


@router.post("/findings/{finding_id}/approve")
async def approve_finding(request: Request, finding_id: str):
    db = get_db(request)
    finding = db.get_finding(finding_id)
    if not finding:
        return JSONResponse({"error": "Finding not found"}, status_code=404)
    if finding["status"] not in ("pending", "rejected"):
        return JSONResponse({"error": f"Cannot approve finding with status '{finding['status']}'"}, status_code=400)
    db.update_finding_status(finding_id, FindingStatus.APPROVED)
    logger.info("Finding %s approved", finding_id)
    return JSONResponse({"ok": True, "status": "approved"})


@router.post("/findings/{finding_id}/reject")
async def reject_finding(request: Request, finding_id: str):
    db = get_db(request)
    finding = db.get_finding(finding_id)
    if not finding:
        return JSONResponse({"error": "Finding not found"}, status_code=404)
    if finding["status"] not in ("pending", "approved"):
        return JSONResponse({"error": f"Cannot reject finding with status '{finding['status']}'"}, status_code=400)
    db.update_finding_status(finding_id, FindingStatus.REJECTED)
    logger.info("Finding %s rejected", finding_id)
    return JSONResponse({"ok": True, "status": "rejected"})


@router.post("/findings/{finding_id}/toggle-test")
async def toggle_finding_test(request: Request, finding_id: str):
    db = get_db(request)
    finding = db.get_finding(finding_id)
    if not finding:
        return JSONResponse({"error": "Finding not found"}, status_code=404)
    new_val = not bool(finding.get("include_test", 1))
    db.set_finding_include_test(finding_id, new_val)
    return JSONResponse({"ok": True, "include_test": new_val})


# --- Scans ---

@router.get("/scans", response_class=HTMLResponse)
async def scans_list(request: Request):
    db = get_db(request)
    scans = db.list_scans()
    log_counts = db.get_scan_log_counts()
    return templates.TemplateResponse(request, "scans.html", {
        "scans": scans,
        "log_counts": log_counts,
    })


# --- System config API ---

@router.get("/api/browse-abs")
async def browse_absolute(path: str = "", file_ext: str = ""):
    """Browse the local filesystem by absolute path.

    When path is empty on Windows, returns the list of available drive letters.
    On Unix, falls through to root.

    If file_ext is provided (e.g. ".uproject"), matching files are included in
    the response alongside directories, with is_file=True.
    """
    import os
    import string

    _skip = {
        "$Recycle.Bin", "System Volume Information", "pagefile.sys",
        "hiberfil.sys", "swapfile.sys", "DumpStack.log.tmp",
        "Config.Msi", "MSOCache", "Recovery",
    }

    def _norm(p: Path) -> str:
        return str(p).replace("\\", "/")

    if not path:
        if os.name == "nt":
            drives = [
                f"{d}:/" for d in string.ascii_uppercase if Path(f"{d}:\\").exists()
            ]
            return JSONResponse({
                "path": "",
                "entries": [{"name": d.rstrip("/"), "path": d} for d in drives],
                "parent": None,
            })
        path = "/"

    target = Path(path).expanduser().resolve()
    if not target.exists() or not target.is_dir():
        return JSONResponse({"error": "Not a directory"}, status_code=404)

    entries = []
    try:
        for entry in sorted(target.iterdir(), key=lambda e: (e.is_file(), e.name.lower())):
            if entry.name in _skip:
                continue
            try:
                if entry.is_dir():
                    entries.append({"name": entry.name, "path": _norm(entry) + "/", "is_file": False})
                elif file_ext and entry.is_file() and entry.suffix.lower() == file_ext.lower():
                    entries.append({"name": entry.name, "path": _norm(entry), "is_file": True})
            except (PermissionError, OSError):
                pass
    except (PermissionError, OSError):
        pass

    # Parent logic: at a drive root (C:/) go back to the drives list ("")
    target_norm = _norm(target)
    parent_norm = _norm(target.parent)
    if os.name == "nt" and parent_norm == target_norm:
        parent = ""          # at drive root → up to drives list
    elif parent_norm == target_norm:
        parent = None        # at filesystem root
    else:
        parent = parent_norm + "/"

    return JSONResponse({
        "path": target_norm + "/",
        "entries": entries,
        "parent": parent,
    })


@router.get("/api/browse")
async def browse_directory(request: Request, path: str = "", base: str = ""):
    """Browse a directory tree.

    ``base`` is an optional absolute path to use as the root instead of the
    configured repo_path.  This is used by the setup wizard when the user is
    configuring a different repo than the currently active one.
    """
    _skip = {"Binaries", "Intermediate", "Saved", "DerivedDataCache", "__pycache__", "node_modules", ".git"}

    if base:
        repo = Path(base).expanduser().resolve()
        if not repo.exists() or not repo.is_dir():
            return JSONResponse({"error": "Base path not found or not a directory"}, status_code=400)
    else:
        config = get_config(request)
        repo = Path(config.repo_path).resolve()

    norm = path.replace("\\", "/").strip("/")
    target = (repo / norm).resolve() if norm else repo
    try:
        target.relative_to(repo)
    except ValueError:
        return JSONResponse({"error": "Invalid path"}, status_code=400)
    if not target.is_dir():
        return JSONResponse({"error": "Not a directory"}, status_code=404)

    entries = []
    try:
        for entry in sorted(target.iterdir(), key=lambda e: e.name.lower()):
            if entry.name.startswith(".") or entry.name in _skip or not entry.is_dir():
                continue
            rel = str(entry.relative_to(repo)).replace("\\", "/")
            entries.append({"name": entry.name, "path": rel + "/"})
    except PermissionError:
        pass

    current_rel = str(target.relative_to(repo)).replace("\\", "/") if target != repo else ""
    parent = None
    if current_rel:
        p = Path(current_rel).parent
        parent = "" if str(p) == "." else str(p).replace("\\", "/") + "/"

    return JSONResponse({"path": current_rel + "/" if current_rel else "", "entries": entries, "parent": parent})


@router.post("/api/open-folder")
async def open_folder(request: Request):
    """Open a local folder in the OS file explorer."""
    import subprocess
    import sys
    data = await request.json()
    path = data.get("path", "").strip()
    if not path:
        return JSONResponse({"error": "No path provided"}, status_code=400)
    p = Path(path)
    if not p.exists():
        return JSONResponse({"error": f"Path does not exist: {path}"}, status_code=404)
    try:
        if sys.platform == "win32":
            subprocess.Popen(["explorer", str(p)])
        elif sys.platform == "darwin":
            subprocess.Popen(["open", str(p)])
        else:
            subprocess.Popen(["xdg-open", str(p)])
        return JSONResponse({"ok": True})
    except Exception as exc:
        return JSONResponse({"error": str(exc)}, status_code=500)


@router.get("/api/systems")
async def get_systems_api(request: Request):
    db = get_db(request)
    return JSONResponse({"systems": db.list_systems()})


@router.get("/api/source-dirs")
async def get_source_dirs_api(request: Request):
    """Return active (non-ignored) source directories from the DB."""
    db = get_db(request)
    dirs = db.list_source_dirs()
    active = [d for d in dirs if d["source_type"] != "ignored"]
    return JSONResponse({"source_dirs": active})


@router.get("/api/source-dirs-all")
async def get_all_source_dirs_api(request: Request):
    """Return all source directories (active + ignored) from the DB."""
    db = get_db(request)
    return JSONResponse({"source_dirs": db.list_source_dirs()})


@router.post("/api/systems/append")
async def append_systems_api(request: Request):
    """Add new systems without touching existing ones."""
    db = get_db(request)
    body = await request.json()
    new_systems = body.get("systems", [])
    for s in new_systems:
        if not s.get("name", "").strip():
            return JSONResponse({"error": "System name cannot be empty"}, status_code=400)
        if not s.get("paths"):
            return JSONResponse({"error": f"System '{s['name']}' has no paths"}, status_code=400)

    existing = db.list_systems()
    existing_names = {s["name"] for s in existing}
    combined = list(existing) + [
        {
            "name": s["name"].strip(),
            "source_dir": s.get("source_dir") or "",
            "paths": s["paths"],
            "min_confidence": s.get("min_confidence") or None,
            "file_extensions": s.get("file_extensions") or None,
            "claude_fast_mode": s.get("claude_fast_mode"),
        }
        for s in new_systems
        if s["name"].strip() not in existing_names
    ]
    db.replace_systems(combined)
    return JSONResponse({"ok": True})


@router.post("/api/systems")
async def save_systems_api(request: Request):
    db = get_db(request)
    body = await request.json()
    systems_data = body.get("systems", [])
    for s in systems_data:
        if not s.get("name", "").strip():
            return JSONResponse({"error": "System name cannot be empty"}, status_code=400)
        if not s.get("paths"):
            return JSONResponse({"error": f"System '{s['name']}' has no paths"}, status_code=400)
    db.replace_systems([
        {
            "name": s["name"].strip(),
            "source_dir": s.get("source_dir") or "",
            "paths": s["paths"],
            "min_confidence": s.get("min_confidence") or None,
            "file_extensions": s.get("file_extensions") or None,
            "claude_fast_mode": s.get("claude_fast_mode"),
        }
        for s in systems_data
    ])
    return JSONResponse({"ok": True})


# --- Project management API ---

@router.get("/api/projects")
async def list_projects(request: Request):
    from auditor.config import list_project_configs
    projects = list_project_configs()
    current_path = getattr(request.app.state, "config_path", "")
    return JSONResponse({"projects": projects, "current": current_path})


@router.post("/api/projects/switch")
async def switch_project(request: Request):
    from auditor.config import load_config, get_db_path
    body = await request.json()
    config_path_str = body.get("path", "").strip()
    if not config_path_str:
        return JSONResponse({"error": "No config path provided"}, status_code=400)
    p = Path(config_path_str)
    if not p.exists():
        return JSONResponse({"error": f"Config file not found: {config_path_str}"}, status_code=404)
    try:
        new_config = load_config(p)
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=400)

    # Switch in-memory config and database
    from auditor.database import Database
    old_db: Database = request.app.state.db
    old_db.close()

    new_db = Database(get_db_path(new_config))
    new_db.init_schema()

    # Migrate systems from YAML if the new config carries them (legacy) and
    # the DB has none, or always restore when the YAML has authoritative data.
    if new_config.systems:
        new_db.replace_systems([
            {
                "name": s.name,
                "source_dir": s.source_dir,
                "paths": list(s.paths),
                "min_confidence": s.min_confidence,
                "file_extensions": list(s.file_extensions) if s.file_extensions else None,
                "claude_fast_mode": s.claude_fast_mode,
            }
            for s in new_config.systems
        ])
        logger.info(
            "Restored %d system(s) from YAML config for project: %s",
            len(new_config.systems), new_config.repo_path,
        )

    request.app.state.config = new_config
    request.app.state.config_path = str(p)
    request.app.state.db = new_db

    from auditor.config import set_active_config_path
    set_active_config_path(p)

    logger.info("Switched to project: %s", new_config.repo_path)
    return JSONResponse({"ok": True, "repo_path": new_config.repo_path})


@router.delete("/api/projects")
async def delete_project(request: Request):
    """Delete the specified project config YAML and its SQLite database.

    If the deleted project is currently active, the server switches to the next
    available project (or a blank state) so it remains operational.
    """
    from auditor.config import (
        AuditorConfig, ACTIVE_POINTER_PATH, list_project_configs,
        set_active_config_path, get_db_path,
    )
    body = await request.json()
    config_path_str = body.get("path", "").strip()
    if not config_path_str:
        return JSONResponse({"error": "No config path provided"}, status_code=400)

    target = Path(config_path_str)
    if not target.exists():
        return JSONResponse({"error": f"Config file not found: {config_path_str}"}, status_code=404)

    current_path = getattr(request.app.state, "config_path", "")
    is_active = str(target) == current_path or target.resolve() == Path(current_path).resolve()

    # Determine the DB path before we delete the YAML
    try:
        from auditor.config import load_config
        target_config = load_config(target)
        target_db_path = get_db_path(target_config)
    except Exception:
        target_db_path = None

    # Close current DB if we're deleting the active project
    if is_active:
        try:
            request.app.state.db.close()
        except Exception:
            pass

    # Delete YAML config
    try:
        target.unlink()
    except Exception as e:
        return JSONResponse({"error": f"Failed to delete config file: {e}"}, status_code=500)

    # Delete associated SQLite database
    if target_db_path:
        db_path = Path(target_db_path)
        for suffix in ["", "-shm", "-wal"]:
            try:
                Path(str(db_path) + suffix).unlink(missing_ok=True)
            except Exception:
                pass

    if is_active:
        # Clear active pointer if it pointed to this file
        try:
            if ACTIVE_POINTER_PATH.exists():
                ACTIVE_POINTER_PATH.unlink()
        except Exception:
            pass

        # Switch to another available project, or start blank
        remaining = [p for p in list_project_configs() if p["path"] != config_path_str]
        if remaining:
            try:
                new_config = load_config(Path(remaining[0]["path"]))
                new_db_path = get_db_path(new_config)
                from auditor.database import Database
                new_db = Database(new_db_path)
                new_db.init_schema()
                request.app.state.config = new_config
                request.app.state.config_path = remaining[0]["path"]
                request.app.state.db = new_db
                set_active_config_path(Path(remaining[0]["path"]))
                logger.info("Deleted project, switched to: %s", new_config.repo_path)
                return JSONResponse({"ok": True, "switched_to": remaining[0]["path"]})
            except Exception as e:
                logger.warning("Could not switch to remaining project after delete: %s", e)

        # No remaining projects — go blank
        from auditor.database import Database
        blank_config = AuditorConfig()
        blank_db = Database(get_db_path(blank_config))
        blank_db.init_schema()
        request.app.state.config = blank_config
        request.app.state.config_path = ""
        request.app.state.db = blank_db
        logger.info("Deleted last project, server is now unconfigured")
        return JSONResponse({"ok": True, "switched_to": None})

    logger.info("Deleted project config: %s", config_path_str)
    return JSONResponse({"ok": True})


def _make_build_config(build_data: dict):
    from auditor.config import BuildConfig
    from pathlib import Path as _Path
    import platform

    ue_dir = build_data.get("ue_installation_dir", "").strip()
    ue_cmd = build_data.get("ue_editor_cmd", "").strip()

    # Auto-derive ue_editor_cmd from the installation directory if not provided
    if ue_dir and not ue_cmd:
        exe = "UnrealEditor-Cmd.exe" if platform.system() == "Windows" else "UnrealEditor-Cmd"
        ue_cmd = str(_Path(ue_dir) / "Engine" / "Binaries" / "Win64" / exe)

    return BuildConfig(
        ue_installation_dir=ue_dir,
        ue_editor_cmd=ue_cmd,
        project_file=build_data.get("project_file", "").strip(),
        build_timeout_seconds=int(build_data.get("build_timeout_seconds", 1800)),
        test_timeout_seconds=int(build_data.get("test_timeout_seconds", 600)),
    )


@router.post("/api/projects/init")
async def init_project(request: Request):
    from auditor.config import (
        AuditorConfig, ScanSchedule, BuildConfig,
        NotificationConfig, save_full_config, DEFAULT_CONFIG_PATH, get_db_path,
    )
    import re
    body = await request.json()
    repo_path = body.get("repo_path", "").strip()
    if not repo_path:
        return JSONResponse({"error": "repo_path is required"}, status_code=400)
    if not Path(repo_path).expanduser().exists():
        return JSONResponse({"error": f"Path does not exist: {repo_path}"}, status_code=400)

    systems_data = body.get("systems", [])
    clean_systems = [
        {
            "name": s["name"].strip(),
            "source_dir": s.get("source_dir", ""),
            "paths": s["paths"],
            "min_confidence": s.get("min_confidence") or None,
            "file_extensions": s.get("file_extensions") or None,
            "claude_fast_mode": s.get("claude_fast_mode"),
        }
        for s in systems_data
        if s.get("name", "").strip()
    ]

    build_data = body.get("build", {})
    schedule_data = body.get("scan_schedule", {})

    config = AuditorConfig(
        repo_path=repo_path,
        build=_make_build_config(build_data),
        scan_schedule=ScanSchedule(
            incremental_interval_hours=int(schedule_data.get("incremental_interval_hours", 4)),
            rotation_enabled=bool(schedule_data.get("rotation_enabled", False)),
            rotation_interval_hours=int(schedule_data.get("rotation_interval_hours", 24)),
        ),
        claude_fast_mode=bool(body.get("claude_fast_mode", True)),
        min_confidence=body.get("min_confidence", "medium"),
    )

    config_path_str = body.get("config_path", "").strip()
    if not config_path_str:
        # Derive from project_name if provided
        project_name = body.get("project_name", "").strip()
        if project_name:
            slug = re.sub(r"[^a-z0-9_-]+", "-", project_name.lower()).strip("-")
            config_path_str = f"~/.code-auditor/{slug}.yaml"
    config_path = Path(config_path_str).expanduser() if config_path_str else DEFAULT_CONFIG_PATH

    try:
        save_full_config(config, config_path)
    except Exception as e:
        return JSONResponse({"error": f"Failed to save config: {e}"}, status_code=500)

    # Save systems and source dir classifications into the DB
    db = get_db(request)

    if clean_systems:
        db.replace_systems(clean_systems)

    source_dirs = body.get("source_dirs", [])
    if source_dirs:
        valid_types = {"active", "ignored"}
        for entry in source_dirs:
            path = (entry.get("path") or "").strip()
            stype = (entry.get("source_type") or "active").strip()
            if stype not in valid_types:
                stype = "active"
            if path:
                db.upsert_source_dir(path, stype)

    from auditor.config import set_active_config_path
    set_active_config_path(config_path)

    logger.info("Project config created at: %s", config_path)
    return JSONResponse({"ok": True, "config_path": str(config_path).replace("\\", "/")})


@router.get("/api/git/branches")
async def git_branches_api(request: Request):
    """Return all local git branches for the configured repo, current branch first."""
    config = get_config(request)
    if not config.repo_path:
        return JSONResponse({"error": "No project configured"}, status_code=400)
    from auditor.pipeline.git_ops import get_local_branches
    branches = get_local_branches(config.repo_path)
    if not branches:
        return JSONResponse({"error": "Could not list git branches — is this a git repo?"}, status_code=400)
    # get_local_branches returns the current branch first, use that as fallback
    configured = config.git_branch or (branches[0] if branches else "")
    return JSONResponse({"branches": branches, "configured": configured})


@router.post("/api/config/branch")
async def set_branch_api(request: Request):
    """Change the configured git branch.

    Wipes all findings + batches + last_scan_commit so the new branch starts
    with a clean slate.  The server does NOT checkout the branch — the user
    controls their working tree.
    """
    from auditor.config import save_full_config
    from auditor.pipeline.git_ops import get_local_branches

    body = await request.json()
    branch = (body.get("branch") or "").strip()
    if not branch:
        return JSONResponse({"error": "branch is required"}, status_code=400)

    config = get_config(request)
    if not config.repo_path:
        return JSONResponse({"error": "No project configured"}, status_code=400)

    # Validate the branch actually exists locally
    branches = get_local_branches(config.repo_path)
    if branch not in branches:
        return JSONResponse(
            {"error": f"Branch '{branch}' not found in local repository"},
            status_code=400,
        )

    db = get_db(request)

    # Wipe findings, batches, and scan baseline — fresh start on the new branch
    wiped = db.wipe_findings()
    db.set_config("last_scan_commit", "")
    logger.info("Branch changed to '%s': wiped %d finding(s) and cleared scan baseline", branch, wiped)

    # Persist the new branch to YAML and update in-memory state
    from pydantic import BaseModel
    new_config = config.model_copy(update={"git_branch": branch})
    config_path_str = getattr(request.app.state, "config_path", "")
    if config_path_str:
        try:
            save_full_config(new_config, Path(config_path_str))
        except Exception as e:
            return JSONResponse({"error": f"Failed to save config: {e}"}, status_code=500)
    request.app.state.config = new_config

    return JSONResponse({"ok": True, "branch": branch, "wiped_findings": wiped})


@router.get("/api/detect-systems")
async def detect_systems_api(request: Request, repo_path: str = ""):
    from auditor.config import detect_systems_from_repo
    if not repo_path:
        config = get_config(request)
        repo_path = config.repo_path
    rp = Path(repo_path).expanduser()
    if not rp.exists():
        return JSONResponse({"error": f"Path does not exist: {repo_path}"}, status_code=400)
    candidates = detect_systems_from_repo(repo_path)
    return JSONResponse({"systems": candidates})


@router.get("/api/find-uproject")
async def find_uproject_api(request: Request, repo_path: str = ""):
    """Return the absolute path to the .uproject file in the given repo, if found."""
    if not repo_path:
        config = get_config(request)
        repo_path = config.repo_path
    rp = Path(repo_path).expanduser() if repo_path else None
    if not rp or not rp.exists():
        return JSONResponse({"error": f"Path does not exist: {repo_path}"}, status_code=400)
    matches = sorted(rp.glob("*.uproject"))
    if not matches:
        return JSONResponse({"uproject": None})
    return JSONResponse({"uproject": str(matches[0]).replace("\\", "/")})


@router.get("/api/detect-source-dirs")
async def detect_source_dirs_api(request: Request, repo_path: str = ""):
    """Return heuristically-classified source directories without touching the DB."""
    from auditor.scanner.source_detector import _heuristic_classify
    if not repo_path:
        config = get_config(request)
        repo_path = config.repo_path
    rp = Path(repo_path).expanduser() if repo_path else None
    if not rp or not rp.exists():
        return JSONResponse({"error": f"Path does not exist: {repo_path}"}, status_code=400)
    classified, unclassified = _heuristic_classify(rp)
    dirs = []
    for path, stype in sorted(classified.items()):
        dirs.append({"path": path, "source_type": stype, "auto": True})
    for path in sorted(unclassified):
        dirs.append({"path": path, "source_type": "project", "auto": False})
    return JSONResponse({"dirs": dirs})


def _build_suggest_systems_prompt(dir_info: dict) -> str:
    import json
    listing = json.dumps(dir_info, indent=2)
    return f"""\
You are helping configure a code analysis tool for an Unreal Engine C++ project.

Below is the structure of each active source directory with its immediate subdirectories and key files.

Your task: for each active source directory, suggest logical "systems" — named groups of sub-paths that Claude should analyse together. Each system belongs to exactly one parent source directory.

Rules:
- Each system MUST include a "source_dir" field matching exactly one of the top-level keys in the listing
- System "paths" must be sub-paths of their parent "source_dir" (or equal to source_dir if no further split is needed)
- Split large source directories by gameplay feature area (Combat, AI, Inventory, UI, etc.)
- Small or plugin directories with few subdirs → one system whose path equals the source_dir
- Use directory paths exactly as shown (with trailing slash)
- System names should be short and descriptive

## Source directory structure

{listing}

## Output Format

```json
{{
  "systems": [
    {{"name": "Combat",   "source_dir": "Source/Game/", "paths": ["Source/Game/Combat/"]}},
    {{"name": "AI",       "source_dir": "Source/Game/", "paths": ["Source/Game/AI/"]}},
    {{"name": "MyPlugin", "source_dir": "Plugins/MyPlugin/", "paths": ["Plugins/MyPlugin/Source/"]}}
  ]
}}
```

Return ONLY the JSON object. No markdown fences, no commentary.\
"""


@router.post("/api/suggest-systems")
async def suggest_systems_api(request: Request):
    """Ask Claude to suggest logical scanning systems based on classified source dirs."""
    body = await request.json()
    repo_path = body.get("repo_path", "")
    source_dirs = body.get("source_dirs", [])  # [{path, source_type}]

    if not repo_path:
        config = get_config(request)
        repo_path = config.repo_path

    repo = Path(repo_path).expanduser()
    if not repo.exists():
        return JSONResponse({"error": f"Path does not exist: {repo_path}"}, status_code=400)

    active_dirs = [d for d in source_dirs if d.get("source_type") != "ignored"]
    if not active_dirs:
        return JSONResponse({"error": "No non-ignored source directories to analyse"}, status_code=400)

    # Build a lightweight directory listing to give Claude context
    _skip = {"Binaries", "Intermediate", "Saved", "DerivedDataCache", "__pycache__", ".git", ".vs", ".idea"}
    _key_exts = {".h", ".cpp", ".cs", ".uplugin", ".uproject"}
    dir_info: dict = {}
    for entry in active_dirs:
        p = entry["path"]
        full = repo / p
        if not full.exists():
            continue
        try:
            subdirs = sorted(
                item.name for item in full.iterdir()
                if item.is_dir() and item.name not in _skip and not item.name.startswith(".")
            )[:20]
            files = sorted(
                item.name for item in full.iterdir()
                if item.is_file() and item.suffix in _key_exts
            )[:10]
            dir_info[p] = {
                "type": entry.get("source_type", "project"),
                "subdirs": subdirs,
                "files": files,
            }
        except OSError:
            dir_info[p] = {"type": entry.get("source_type", "project"), "subdirs": [], "files": []}

    if not dir_info:
        return JSONResponse({"error": "Could not read source directories"}, status_code=400)

    prompt = _build_suggest_systems_prompt(dir_info)

    try:
        import subprocess as _sp
        from auditor.analysis.engine import call_claude, _extract_json
        try:
            raw = call_claude(prompt, fast=False, timeout=90, repo_path=repo_path, use_tools=False)
        except _sp.CalledProcessError as cpe:
            # Surface Claude's stderr so the user can diagnose the failure
            stderr_detail = (cpe.stderr or "").strip() or (cpe.stdout or "").strip()
            error_msg = stderr_detail or f"Claude CLI exited with code {cpe.returncode}"
            logger.error("suggest-systems: Claude CLI error (code %d): %s", cpe.returncode, error_msg)
            return JSONResponse({"error": f"Claude CLI error: {error_msg}"}, status_code=500)
        except FileNotFoundError:
            return JSONResponse({"error": "Claude CLI not found — ensure 'claude' is on PATH"}, status_code=500)
        data = _extract_json(raw)
        systems = data.get("systems", [])
        active_dir_paths = {d["path"] for d in active_dirs}
        valid = []
        for s in systems:
            name = (s.get("name") or "").strip()
            source_dir = (s.get("source_dir") or "").strip()
            paths = [p for p in (s.get("paths") or []) if isinstance(p, str) and p.strip()]
            if name and paths:
                # If Claude returned an unknown source_dir, infer from the first path
                if source_dir not in active_dir_paths:
                    source_dir = next(
                        (d for d in active_dir_paths if paths[0].startswith(d)), ""
                    )
                valid.append({"name": name, "source_dir": source_dir, "paths": paths})
        return JSONResponse({"systems": valid})
    except Exception as e:
        logger.exception("suggest-systems failed")
        return JSONResponse({"error": str(e)}, status_code=500)


def _build_suggest_paths_prompt(system_name: str, source_dir: str, subdirs: list[dict]) -> str:
    lines = []
    for d in subdirs:
        lines.append(f"  {d['path']}")
        for c in d["children"][:12]:
            lines.append(f"    {c}")
    listing = "\n".join(lines) or f"  (no subdirectories — source dir is a leaf)"
    sd_root = source_dir.rstrip("/") + "/"
    return f"""\
You are configuring a code analysis tool for an Unreal Engine C++ project.

System name: "{system_name}"
Source directory: "{source_dir}"

Subdirectories of "{source_dir}":
{listing}

Which paths should the system "{system_name}" scan? Choose directories whose names suggest they \
are related to the system name. When in doubt, prefer fewer, broader paths.

Rules:
- Return only paths that appear in the listing above
- If no subdirectory is clearly related, return the entire source dir: "{sd_root}"
- All paths must have a trailing slash
- Do not invent paths that are not listed

Return ONLY valid JSON (no fences, no commentary):
{{"paths": ["{sd_root}"]}}"""


@router.post("/api/suggest-paths")
async def suggest_paths_api(request: Request):
    """Ask Claude to suggest scan paths for a single named system inside its source_dir."""
    body = await request.json()
    system_name = (body.get("system_name") or "").strip()
    source_dir  = (body.get("source_dir")  or "").strip()

    if not system_name:
        return JSONResponse({"error": "system_name is required"}, status_code=400)
    if not source_dir:
        return JSONResponse({"error": "source_dir is required"}, status_code=400)

    config = get_config(request)
    repo_path = (body.get("repo_path") or config.repo_path or "").strip()
    if not repo_path:
        return JSONResponse({"error": "No repo path configured"}, status_code=400)

    repo = Path(repo_path).expanduser().resolve()
    sd_norm  = source_dir.replace("\\", "/").rstrip("/")
    sd_path  = (repo / sd_norm).resolve()
    sd_root  = sd_norm + "/"

    if not sd_path.exists() or not sd_path.is_dir():
        return JSONResponse({"error": f"Source directory not found: {source_dir}"}, status_code=400)

    _skip = {"Binaries", "Intermediate", "Saved", "DerivedDataCache", "__pycache__", ".git", ".vs", ".idea"}

    # Build one-level-deep listing with immediate children for context
    subdirs: list[dict] = []
    try:
        for entry in sorted(sd_path.iterdir(), key=lambda e: e.name.lower()):
            if entry.name in _skip or entry.name.startswith(".") or not entry.is_dir():
                continue
            rel = str(entry.relative_to(repo)).replace("\\", "/").rstrip("/") + "/"
            children: list[str] = []
            try:
                for child in sorted(entry.iterdir(), key=lambda e: e.name.lower()):
                    if child.name in _skip or child.name.startswith(".") or not child.is_dir():
                        continue
                    children.append(str(child.relative_to(repo)).replace("\\", "/").rstrip("/") + "/")
            except (PermissionError, OSError):
                pass
            subdirs.append({"path": rel, "children": children})
    except (PermissionError, OSError) as exc:
        return JSONResponse({"error": f"Cannot read source directory: {exc}"}, status_code=500)

    # No subdirs → the whole source_dir is the only sensible path
    if not subdirs:
        return JSONResponse({"paths": [sd_root]})

    prompt = _build_suggest_paths_prompt(system_name, sd_root, subdirs)
    try:
        import subprocess as _sp
        from auditor.analysis.engine import call_claude, _extract_json
        try:
            raw = call_claude(prompt, fast=True, timeout=60, repo_path=repo_path, use_tools=False)
        except _sp.CalledProcessError as cpe:
            stderr_detail = (cpe.stderr or "").strip() or (cpe.stdout or "").strip()
            return JSONResponse({"error": f"Claude CLI error: {stderr_detail or cpe.returncode}"}, status_code=500)
        except FileNotFoundError:
            return JSONResponse({"error": "Claude CLI not found — ensure 'claude' is on PATH"}, status_code=500)

        data = _extract_json(raw)
        raw_paths = [p for p in (data.get("paths") or []) if isinstance(p, str) and p.strip()]

        # Validate: every path must be under the source_dir
        valid = [p for p in raw_paths if p.replace("\\", "/").startswith(sd_root) or p.replace("\\", "/").rstrip("/") + "/" == sd_root]
        if not valid:
            valid = [sd_root]
        return JSONResponse({"paths": valid})
    except Exception:
        logger.exception("suggest-paths failed")
        return JSONResponse({"error": "Suggestion failed — check server logs"}, status_code=500)


@router.get("/api/config/status")
async def config_status(request: Request):
    from auditor.config import validate_config_errors, get_db_path, DEFAULT_CONFIG_PATH
    config = get_config(request)
    db = get_db(request)

    errors = validate_config_errors(config, systems=db.list_systems())
    db_path = get_db_path(config)
    db_size = db_path.stat().st_size if db_path.exists() else 0
    last_commit = db.get_config("last_scan_commit", "")

    repo = Path(config.repo_path).expanduser()
    config_path = getattr(request.app.state, "config_path", str(DEFAULT_CONFIG_PATH))
    return JSONResponse({
        "config_path": config_path,
        "repo_path": config.repo_path,
        "repo_exists": repo.exists(),
        "errors": errors,
        "last_commit": last_commit,
        "db_size_bytes": db_size,
    })


@router.post("/api/config/repair")
async def repair_config(request: Request):
    """Re-save the active config with all Pydantic defaults filled in."""
    from auditor.config import save_full_config, DEFAULT_CONFIG_PATH
    config = get_config(request)
    config_path = Path(getattr(request.app.state, "config_path", str(DEFAULT_CONFIG_PATH)))
    try:
        save_full_config(config, config_path)
        logger.info("Config repaired at: %s", config_path)
        return JSONResponse({"ok": True})
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


@router.post("/api/config/update")
async def update_config_api(request: Request):
    """Update build, schedule and quality settings in the active config file."""
    from auditor.config import save_full_config, ScanSchedule, AuditorConfig

    body = await request.json()
    config = get_config(request)
    config_path_str = getattr(request.app.state, "config_path", None)
    if not config_path_str:
        return JSONResponse(
            {"error": "No config file loaded — create a project first"},
            status_code=400,
        )

    build_data = body.get("build", {})
    sched_data = body.get("scan_schedule", {})

    new_config = AuditorConfig(
        repo_path=config.repo_path,
        data_dir=config.data_dir,
        notifications=config.notifications,
        build=_make_build_config({
            "ue_installation_dir": build_data.get("ue_installation_dir", config.build.ue_installation_dir),
            "project_file":        build_data.get("project_file",        config.build.project_file),
            "build_timeout_seconds": build_data.get("build_timeout_seconds", config.build.build_timeout_seconds),
            "test_timeout_seconds":  build_data.get("test_timeout_seconds",  config.build.test_timeout_seconds),
        }),
        scan_schedule=ScanSchedule(
            incremental_interval_hours=int(sched_data.get(
                "incremental_interval_hours", config.scan_schedule.incremental_interval_hours)),
            rotation_enabled=bool(sched_data.get(
                "rotation_enabled", config.scan_schedule.rotation_enabled)),
            rotation_interval_hours=int(sched_data.get(
                "rotation_interval_hours", config.scan_schedule.rotation_interval_hours)),
        ),
        claude_fast_mode=bool(body.get("claude_fast_mode", config.claude_fast_mode)),
        min_confidence=body.get("min_confidence", config.min_confidence),
        file_extensions=body.get("file_extensions", list(config.file_extensions)),
    )

    try:
        save_full_config(new_config, Path(config_path_str))
        request.app.state.config = new_config
        logger.info("Config updated at: %s", config_path_str)
        return JSONResponse({"ok": True})
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


@router.post("/scans/trigger")
async def trigger_scan(request: Request):
    config = get_config(request)
    db = get_db(request)
    db_path = db.db_path

    body = await request.json() if request.headers.get("content-type", "").startswith("application/json") else {}
    scan_type = body.get("scan_type", "full")

    # Accept system_id (preferred) or legacy system_name
    system_name: str | None = None
    system_id = body.get("system_id")
    if system_id is not None:
        matched = next((s for s in db.list_systems() if s["id"] == system_id), None)
        if matched is None:
            return JSONResponse({"error": f"System id {system_id!r} not found"}, status_code=404)
        system_name = matched["name"]
    else:
        system_name = body.get("system_name") or None

    # Reject if a scan is already running
    running = db.get_running_scan()
    if running:
        return JSONResponse(
            {"error": "A scan is already running", "scan_id": running["id"]},
            status_code=409,
        )

    from auditor.scan_state import canceller
    canceller.reset()

    def _run():
        from auditor.database import Database
        from auditor.scanner.scheduler import run_scan
        thread_db = Database(db_path)
        try:
            scan_id = run_scan(config, thread_db, scan_type=scan_type, system_name=system_name)
            logger.info("Scan completed: %s", scan_id)
        except Exception:
            logger.exception("Scan failed")
        finally:
            thread_db.close()

    thread = threading.Thread(target=_run, daemon=True)
    thread.start()
    return JSONResponse({"ok": True})


@router.delete("/scans/{scan_id}")
async def delete_scan(request: Request, scan_id: str):
    db = get_db(request)
    scan = db.get_scan(scan_id)
    if not scan:
        return JSONResponse({"error": "Scan not found"}, status_code=404)
    if scan["status"] == "running":
        return JSONResponse({"error": "Cannot delete a running scan"}, status_code=400)
    db.delete_scan(scan_id)
    logger.info("Scan %s deleted", scan_id)
    return JSONResponse({"ok": True})


@router.post("/scans/cancel")
async def cancel_scan(request: Request):
    from auditor.scan_state import canceller
    db = get_db(request)

    if not canceller.is_cancelled:
        canceller.cancel()
        logger.info("Scan cancellation requested")

    from auditor.ws_manager import manager as ws_manager
    running = db.get_running_scan()
    ws_manager.push_scan_status(running=running is not None, scan=running, cancelling=True)

    if running:
        from auditor.models import ScanStatus, now_iso
        db.update_scan(running["id"], status=ScanStatus.CANCELLED, completed_at=now_iso())

    return JSONResponse({"ok": True})


# --- Settings ---

@router.get("/settings", response_class=HTMLResponse)
async def settings_page(request: Request):
    from auditor.config import DEFAULT_CONFIG_PATH
    db = get_db(request)
    config = get_config(request)
    source_dirs = db.list_source_dirs()
    active_dirs = [d for d in source_dirs if d["source_type"] != "ignored"]
    ignored_dirs = [d for d in source_dirs if d["source_type"] == "ignored"]

    # Count systems per source_dir for display
    all_systems = db.list_systems()
    sys_count: dict[str, int] = {}
    for s in all_systems:
        sd = (s.get("source_dir") or "").strip()
        sys_count[sd] = sys_count.get(sd, 0) + 1
    for d in active_dirs:
        d["system_count"] = sys_count.get(d["path"], 0)

    config_path = getattr(request.app.state, "config_path", "") or ""

    # Show whatever branch is stored in config — no subprocess on page load.
    configured_branch = config.git_branch or ""

    return templates.TemplateResponse(request, "settings.html", {
        "active_dirs": active_dirs,
        "ignored_dirs": ignored_dirs,
        "config": config,
        "config_path": config_path,
        "configured_branch": configured_branch,
    })


@router.post("/settings/source-dirs")
async def update_source_dir(request: Request):
    db = get_db(request)
    body = await request.json()
    path = body.get("path", "").strip()
    source_type = body.get("source_type", "").strip()
    if not path or source_type not in ("active", "ignored"):
        return JSONResponse({"error": "Invalid path or source_type"}, status_code=400)
    db.upsert_source_dir(path, source_type)
    logger.info("Source dir updated: '%s' -> '%s'", path, source_type)
    return JSONResponse({"ok": True, "path": path, "source_type": source_type})


@router.delete("/settings/source-dirs")
async def delete_source_dir(request: Request):
    db = get_db(request)
    body = await request.json()
    path = body.get("path", "").strip()
    if not path:
        return JSONResponse({"error": "Invalid path"}, status_code=400)
    db.delete_source_dir(path)
    logger.info("Source dir deleted: '%s'", path)
    return JSONResponse({"ok": True, "path": path})


@router.post("/settings/source-dirs/bulk")
async def bulk_update_source_dirs(request: Request):
    """Apply a batch of upserts and deletes to source_dirs in one request.

    Body: {
        "upsert": [{"path": "...", "source_type": "active"|"ignored"}, ...],
        "delete": ["path1", "path2", ...]
    }
    """
    db = get_db(request)
    body = await request.json()
    valid_types = {"active", "ignored"}

    upserted = []
    for entry in body.get("upsert", []):
        path = (entry.get("path") or "").strip()
        stype = (entry.get("source_type") or "active").strip()
        if not path or stype not in valid_types:
            continue
        db.upsert_source_dir(path, stype)
        upserted.append(path)

    deleted = []
    for path in body.get("delete", []):
        path = (path or "").strip()
        if not path:
            continue
        db.delete_source_dir(path)
        deleted.append(path)

    logger.info("Bulk source-dirs: upserted %d, deleted %d", len(upserted), len(deleted))
    return JSONResponse({"ok": True, "upserted": upserted, "deleted": deleted})


# --- Batches ---

@router.get("/batches", response_class=HTMLResponse)
async def batches_list(request: Request):
    db = get_db(request)
    batches = db.list_batches()
    return templates.TemplateResponse(request, "batches.html", {
        "batches": batches,
    })


@router.get("/batches/{batch_id}", response_class=HTMLResponse)
async def batch_detail(request: Request, batch_id: str):
    db = get_db(request)
    batch = db.get_batch(batch_id)
    if not batch:
        return HTMLResponse("<h1>Batch not found</h1>", status_code=404)
    findings = [db.get_finding(fid) for fid in batch["finding_ids"]]
    findings = [f for f in findings if f]
    return templates.TemplateResponse(request, "batch_status.html", {
        "batch": batch,
        "findings": findings,
    })


@router.post("/batch/apply")
async def apply_batch(request: Request):
    config = get_config(request)
    db = get_db(request)

    approved = db.get_approved_findings()
    if not approved:
        return JSONResponse({"error": "No approved findings to apply"}, status_code=400)

    from auditor.models import Batch, new_id
    batch = Batch(
        id=new_id(),
        finding_ids=[f["id"] for f in approved],
    )
    db.insert_batch(batch)

    for f in approved:
        db.set_finding_batch(f["id"], batch.id)

    db_path = db.db_path
    batch_id = batch.id

    def _run():
        from auditor.database import Database
        from auditor.pipeline.batch import run_batch_pipeline
        thread_db = Database(db_path)
        try:
            run_batch_pipeline(config, thread_db, batch_id)
        except Exception:
            logger.exception("Batch pipeline failed for %s", batch_id)
            thread_db.update_batch(batch_id, status=BatchStatus.FAILED)
        finally:
            thread_db.close()

    thread = threading.Thread(target=_run, daemon=True)
    thread.start()

    return JSONResponse({"ok": True, "batch_id": batch.id})


# --- API ---

@router.get("/api/stats")
async def api_stats(request: Request):
    db = get_db(request)
    return JSONResponse(db.get_stats())


@router.websocket("/ws")
async def websocket_endpoint(websocket: WebSocket):
    from auditor.ws_manager import manager as ws_manager
    from auditor.scan_state import canceller

    await ws_manager.connect(websocket)
    try:
        db = websocket.app.state.db
        running = db.get_running_scan()
        await websocket.send_text(json.dumps({
            "type": "scan_status",
            "running": running is not None,
            "scan": running,
            "cancelling": canceller.is_cancelled,
        }))
        while True:
            await websocket.receive_text()
    except WebSocketDisconnect:
        pass
    except Exception:
        pass
    finally:
        ws_manager.disconnect(websocket)


@router.get("/api/scans/{scan_id}/logs")
async def api_scan_logs(request: Request, scan_id: str, offset: int = 0):
    db = get_db(request)
    logs = db.get_scan_logs(scan_id, offset=offset)
    scan = db.get_scan(scan_id)
    running = scan["status"] == "running" if scan else False
    return JSONResponse({"logs": logs, "running": running, "total": len(logs)})


@router.get("/api/findings/stream")
async def api_findings_stream(request: Request, scan_id: str, offset: int = 0):
    """Return findings for a scan starting from offset N, respecting active filters."""
    db = get_db(request)
    findings = db.get_scan_findings_from(scan_id, offset)
    return JSONResponse({"findings": findings, "total": len(findings)})


@router.get("/api/scan-status")
async def api_scan_status(request: Request):
    from auditor.scan_state import canceller
    db = get_db(request)
    running = db.get_running_scan()
    return JSONResponse({
        "running": running is not None,
        "scan": running,
        "cancelling": canceller.is_cancelled,
    })
